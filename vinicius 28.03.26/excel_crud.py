from datetime import datetime
from pathlib import Path
import json
import math
import os
import random
import re
import sys
import tkinter as tk
from tkinter import messagebox, ttk

from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook

BASE_RECURSOS = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
BASE_DADOS = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent

ARQUIVO_EXCEL = BASE_DADOS / "clientes.xlsx"
ARQUIVO_LOGO = BASE_RECURSOS / "logo_transparente.png"
ARQUIVO_CONFIG = BASE_DADOS / "config.json"
EXTENSOES_IMAGEM = [".png", ".jpg", ".jpeg", ".webp"]
ABA_CLIENTES = "clientes"
ABA_PRODUTOS = "produtos"
ABA_VENDAS = "vendas"
CABECALHO = [
    "id",
    "nome",
    "cpf",
    "email",
    "telefone",
    "idade",
    "cidade",
    "estado",
    "status",
    "data_cadastro",
]
CABECALHO_PRODUTOS = ["id", "nome", "categoria", "preco", "estoque", "status", "data_cadastro"]
CABECALHO_VENDAS = [
    "id",
    "produto_id",
    "produto_nome",
    "cliente",
    "quantidade",
    "valor_unitario",
    "valor_total",
    "status",
    "data_venda",
]
STATUS_PADRAO = "Ativo"
CAMPOS_FORMULARIO = ["nome", "cpf", "email", "telefone", "idade", "cidade", "estado", "status"]
STATUS_PRODUTO_PADRAO = "Disponivel"
PADRAO_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.com$")
FUNDO_JANELA = "#05010a"
FUNDO_CARD = "#0d0d12"
FUNDO_CAMPO = "#d9dbe1"
FUNDO_DESTAQUE = "#7b2cbf"
AZUL_ACAO = "#c77dff"
AZUL_SUAVE = "#f8f9fa"
VERDE_ACAO = "#9d4edd"
VERMELHO_ACAO = "#ff4d8d"
TEXTO_CLARO = "#f8f9ff"
TEXTO_MUTED = "#d6c7f5"
ROXO_NEON = "#c77dff"
ROXO_NEON_FORTE = "#e0aaff"
FUNDO_TABELA = "#d6d8df"
TEXTO_ESCURO = "#111827"
EMAIL_DONO_PADRAO = "admin@orquideas.local"
SENHA_DONO_PADRAO = "1234"
CHAVE_PIX_PADRAO = "pix@orquideas.local"


def carregar_configuracoes():
    configuracoes_padrao = {
        "admin_email": EMAIL_DONO_PADRAO,
        "admin_password": SENHA_DONO_PADRAO,
        "pix_key": CHAVE_PIX_PADRAO,
    }

    if ARQUIVO_CONFIG.exists():
        try:
            with ARQUIVO_CONFIG.open("r", encoding="utf-8") as arquivo:
                conteudo = json.load(arquivo)
            if isinstance(conteudo, dict):
                configuracoes_padrao.update(
                    {
                        "admin_email": str(conteudo.get("admin_email") or configuracoes_padrao["admin_email"]).strip(),
                        "admin_password": str(conteudo.get("admin_password") or configuracoes_padrao["admin_password"]),
                        "pix_key": str(conteudo.get("pix_key") or configuracoes_padrao["pix_key"]).strip(),
                    }
                )
        except (OSError, json.JSONDecodeError):
            pass

    configuracoes_padrao["admin_email"] = os.getenv("ORQ_ADMIN_EMAIL", configuracoes_padrao["admin_email"]).strip()
    configuracoes_padrao["admin_password"] = os.getenv("ORQ_ADMIN_PASSWORD", configuracoes_padrao["admin_password"])
    configuracoes_padrao["pix_key"] = os.getenv("ORQ_PIX_KEY", configuracoes_padrao["pix_key"]).strip()
    return configuracoes_padrao


CONFIG = carregar_configuracoes()
EMAIL_DONO = CONFIG["admin_email"]
SENHA_DONO = CONFIG["admin_password"]
CHAVE_PIX = CONFIG["pix_key"]


class FundoAnimado:
    def __init__(self, raiz, quantidade=10, margem=220):
        self.raiz = raiz
        self.canvas = tk.Canvas(raiz, bg=FUNDO_JANELA, highlightthickness=0, bd=0)
        self.canvas.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.quantidade = quantidade
        self.margem = margem
        self.fase = 0.0
        self.linhas = []
        self.ultima_largura = 0
        self.ultima_altura = 0
        self.raiz.bind("<Configure>", self.redesenhar)
        self.criar_raizes()
        self.animar()
        self.enviar_para_fundo()

    def criar_raizes(self):
        self.linhas.clear()
        largura = max(self.raiz.winfo_width(), 800)
        altura = max(self.raiz.winfo_height(), 500)
        self.ultima_largura = largura
        self.ultima_altura = altura
        recuo = 24
        configuracoes = [
            ("topo", largura, recuo),
            ("base", largura, altura - recuo),
            ("esquerda", altura, recuo),
            ("direita", altura, largura - recuo),
        ]

        for lado, extensao, fixa in configuracoes:
            for _ in range(max(2, self.quantidade // 4)):
                self.linhas.append(
                    {
                        "lado": lado,
                        "extensao": extensao,
                        "fixa": fixa,
                        "offset": random.randint(-120, 120),
                        "amplitude": random.randint(12, 28),
                        "frequencia": random.uniform(0.012, 0.026),
                        "velocidade": random.uniform(0.9, 1.8),
                        "espessura": random.randint(5, 8),
                        "fase_local": random.uniform(0, math.pi * 2),
                        "espinhos": random.randint(26, 42),
                    }
                )

    def desenhar_espinhos(self, pontos, lado, quantidade, espessura):
        total_pontos = len(pontos) // 2
        if total_pontos < 6:
            return

        passo = max(2, total_pontos // quantidade)
        for indice_ponto in range(2, total_pontos - 2, passo):
            x = pontos[indice_ponto * 2]
            y = pontos[indice_ponto * 2 + 1]

            if lado == "topo":
                ponta_x, ponta_y = x, y + random.randint(8, 18)
            elif lado == "base":
                ponta_x, ponta_y = x, y - random.randint(8, 18)
            elif lado == "esquerda":
                ponta_x, ponta_y = x + random.randint(8, 18), y
            else:
                ponta_x, ponta_y = x - random.randint(8, 18), y

            self.canvas.create_line(
                x,
                y,
                ponta_x,
                ponta_y,
                fill=ROXO_NEON_FORTE,
                width=max(2, espessura // 3),
                tags="raiz",
            )

    def desenhar(self):
        self.canvas.delete("raiz")
        largura = max(self.raiz.winfo_width(), 800)
        altura = max(self.raiz.winfo_height(), 500)
        self.canvas.configure(width=largura, height=altura)

        for linha in self.linhas:
            pontos = []
            segmentos = 28
            for indice in range(segmentos + 1):
                progresso = indice / segmentos
                base = progresso * linha["extensao"] + linha["offset"]
                fase_onda = self.fase * linha["velocidade"] + base * linha["frequencia"] + linha["fase_local"]
                onda = math.sin(fase_onda) * linha["amplitude"]

                if linha["lado"] == "topo":
                    x = base
                    y = linha["fixa"] + onda
                elif linha["lado"] == "base":
                    x = base
                    y = linha["fixa"] - onda
                elif linha["lado"] == "esquerda":
                    x = linha["fixa"] + onda
                    y = base
                else:
                    x = linha["fixa"] - onda
                    y = base

                pontos.extend([x, y])

            self.canvas.create_line(
                pontos,
                fill="#4b145f",
                width=linha["espessura"] + 8,
                smooth=True,
                splinesteps=36,
                tags="raiz",
            )
            self.canvas.create_line(
                pontos,
                fill=ROXO_NEON,
                width=linha["espessura"] + 3,
                smooth=True,
                splinesteps=36,
                tags="raiz",
            )
            self.canvas.create_line(
                pontos,
                fill=ROXO_NEON_FORTE,
                width=max(2, linha["espessura"] - 1),
                smooth=True,
                splinesteps=36,
                tags="raiz",
            )
            self.desenhar_espinhos(pontos, linha["lado"], linha["espinhos"], linha["espessura"])

    def animar(self):
        self.fase += 0.06
        self.desenhar()
        self.raiz.after(45, self.animar)

    def redesenhar(self, _evento=None):
        largura = max(self.raiz.winfo_width(), 800)
        altura = max(self.raiz.winfo_height(), 500)

        precisa_recriar = not self.linhas
        precisa_recriar = precisa_recriar or abs(largura - self.ultima_largura) > 120
        precisa_recriar = precisa_recriar or abs(altura - self.ultima_altura) > 120

        if precisa_recriar:
            self.criar_raizes()
        self.desenhar()

    def enviar_para_fundo(self):
        self.canvas.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.canvas.tk.call("lower", self.canvas._w)


def inicializar_planilha():
    if not ARQUIVO_EXCEL.exists():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = ABA_CLIENTES
        worksheet.append(CABECALHO)
        worksheet_produtos = workbook.create_sheet(ABA_PRODUTOS)
        worksheet_produtos.append(CABECALHO_PRODUTOS)
        worksheet_vendas = workbook.create_sheet(ABA_VENDAS)
        worksheet_vendas.append(CABECALHO_VENDAS)
        workbook.save(ARQUIVO_EXCEL)

    migrar_planilha_se_necessario()
    garantir_abas_extras()


def abrir_planilha():
    workbook = load_workbook(ARQUIVO_EXCEL)
    worksheet = workbook[ABA_CLIENTES]
    return workbook, worksheet


def garantir_abas_extras():
    workbook = load_workbook(ARQUIVO_EXCEL)

    if ABA_PRODUTOS not in workbook.sheetnames:
        worksheet_produtos = workbook.create_sheet(ABA_PRODUTOS)
        worksheet_produtos.append(CABECALHO_PRODUTOS)

    if ABA_VENDAS not in workbook.sheetnames:
        worksheet_vendas = workbook.create_sheet(ABA_VENDAS)
        worksheet_vendas.append(CABECALHO_VENDAS)

    workbook.save(ARQUIVO_EXCEL)
    workbook.close()


def migrar_planilha_se_necessario():
    workbook, worksheet = abrir_planilha()
    cabecalho_atual = [celula.value for celula in worksheet[1]]

    if cabecalho_atual == CABECALHO:
        workbook.close()
        return

    registros_antigos = []
    indices = {nome_coluna: indice for indice, nome_coluna in enumerate(cabecalho_atual)}
    formato_antigo_basico = cabecalho_atual == ["id", "nome", "email", "idade"]

    def valor_coluna(linha, nome_coluna, padrao=""):
        indice = indices.get(nome_coluna)
        if indice is None or indice >= len(linha):
            return padrao
        valor = linha[indice]
        return padrao if valor is None else valor

    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue

        if formato_antigo_basico:
            registros_antigos.append(
                {
                    "id": linha[0],
                    "nome": linha[1] if len(linha) > 1 and linha[1] is not None else "",
                    "cpf": "",
                    "email": linha[2] if len(linha) > 2 and linha[2] is not None else "",
                    "telefone": "",
                    "idade": linha[3] if len(linha) > 3 and linha[3] is not None else "",
                    "cidade": "",
                    "estado": "",
                    "status": STATUS_PADRAO,
                    "data_cadastro": datetime.now().strftime("%d/%m/%Y %H:%M"),
                }
            )
            continue

        registros_antigos.append(
            {
                "id": linha[0],
                "nome": valor_coluna(linha, "nome"),
                "cpf": valor_coluna(linha, "cpf"),
                "email": valor_coluna(linha, "email"),
                "telefone": valor_coluna(linha, "telefone"),
                "idade": valor_coluna(linha, "idade"),
                "cidade": valor_coluna(linha, "cidade"),
                "estado": valor_coluna(linha, "estado"),
                "status": valor_coluna(linha, "status", STATUS_PADRAO),
                "data_cadastro": valor_coluna(
                    linha, "data_cadastro", datetime.now().strftime("%d/%m/%Y %H:%M")
                ),
            }
        )

    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(CABECALHO)

    for cliente in registros_antigos:
        worksheet.append(
            [
                cliente["id"],
                cliente["nome"],
                cliente["cpf"],
                cliente["email"],
                cliente["telefone"],
                cliente["idade"],
                cliente["cidade"],
                cliente["estado"],
                cliente["status"] or STATUS_PADRAO,
                cliente["data_cadastro"],
            ]
        )

    workbook.save(ARQUIVO_EXCEL)
    workbook.close()


def proximo_id(worksheet):
    ids = []
    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is not None:
            ids.append(int(linha[0]))
    return max(ids, default=0) + 1


def normalizar_cpf(cpf):
    return re.sub(r"\D", "", str(cpf or ""))


def email_aceito(email):
    if not PADRAO_EMAIL.match(email):
        return False
    dominio = email.rsplit("@", 1)[-1]
    return "." in dominio


def normalizar_cliente(linha):
    valores = list(linha) + [""] * (len(CABECALHO) - len(linha))
    return {coluna: "" if valores[indice] is None else str(valores[indice]) for indice, coluna in enumerate(CABECALHO)}


def listar_clientes():
    workbook, worksheet = abrir_planilha()
    clientes = []

    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue
        clientes.append(normalizar_cliente(linha))

    workbook.close()
    return clientes


def buscar_cliente_por_id(cliente_id):
    workbook, worksheet = abrir_planilha()

    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] == cliente_id:
            cliente = normalizar_cliente(linha)
            workbook.close()
            return cliente

    workbook.close()
    return None


def buscar_cliente_por_cpf(cpf):
    workbook, worksheet = abrir_planilha()
    cpf_normalizado = normalizar_cpf(cpf)

    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue
        if normalizar_cpf(linha[2]) == cpf_normalizado:
            cliente = normalizar_cliente(linha)
            workbook.close()
            return cliente

    workbook.close()
    return None


def buscar_cliente_por_nome_cpf(nome, cpf):
    nome_normalizado = nome.strip().lower()
    cpf_normalizado = normalizar_cpf(cpf)
    for cliente in listar_clientes():
        if cliente["nome"].strip().lower() == nome_normalizado and normalizar_cpf(cliente["cpf"]) == cpf_normalizado:
            return cliente
    return None


def cpf_ja_cadastrado(cpf, cliente_id_atual=None):
    cpf_normalizado = normalizar_cpf(cpf)
    if not cpf_normalizado:
        return False

    for cliente in listar_clientes():
        if normalizar_cpf(cliente["cpf"]) != cpf_normalizado:
            continue
        if cliente_id_atual is not None and int(cliente["id"]) == int(cliente_id_atual):
            continue
        return True
    return False


def buscar_clientes_por_termo(termo):
    termo_normalizado = termo.strip().lower()
    if not termo_normalizado:
        return listar_clientes()

    clientes_filtrados = []
    for cliente in listar_clientes():
        if any(termo_normalizado in str(valor).lower() for valor in cliente.values()):
            clientes_filtrados.append(cliente)
    return clientes_filtrados


def abrir_aba(nome_aba):
    workbook = load_workbook(ARQUIVO_EXCEL)
    worksheet = workbook[nome_aba]
    return workbook, worksheet


def proximo_id_generico(worksheet):
    ids = []
    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is not None:
            ids.append(int(linha[0]))
    return max(ids, default=0) + 1


def normalizar_registro(linha, cabecalho):
    valores = list(linha) + [""] * (len(cabecalho) - len(linha))
    return {coluna: "" if valores[indice] is None else str(valores[indice]) for indice, coluna in enumerate(cabecalho)}


def listar_produtos():
    workbook, worksheet = abrir_aba(ABA_PRODUTOS)
    produtos = []
    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue
        produtos.append(normalizar_registro(linha, CABECALHO_PRODUTOS))
    workbook.close()
    return produtos


def buscar_imagem_produto(nome_produto):
    for extensao in EXTENSOES_IMAGEM:
        caminho = BASE_RECURSOS / f"{nome_produto}{extensao}"
        if caminho.exists():
            return caminho
    return None


def criar_produto(dados):
    workbook, worksheet = abrir_aba(ABA_PRODUTOS)
    novo_id = proximo_id_generico(worksheet)
    data_cadastro = datetime.now().strftime("%d/%m/%Y %H:%M")
    worksheet.append(
        [
            novo_id,
            dados["nome"],
            dados["categoria"],
            dados["preco"],
            dados["estoque"],
            dados["status"] or STATUS_PRODUTO_PADRAO,
            data_cadastro,
        ]
    )
    workbook.save(ARQUIVO_EXCEL)
    workbook.close()


def atualizar_produto(produto_id, dados):
    workbook, worksheet = abrir_aba(ABA_PRODUTOS)
    for linha in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=linha, column=1).value == produto_id:
            worksheet.cell(row=linha, column=2, value=dados["nome"])
            worksheet.cell(row=linha, column=3, value=dados["categoria"])
            worksheet.cell(row=linha, column=4, value=dados["preco"])
            worksheet.cell(row=linha, column=5, value=dados["estoque"])
            worksheet.cell(row=linha, column=6, value=dados["status"] or STATUS_PRODUTO_PADRAO)
            workbook.save(ARQUIVO_EXCEL)
            workbook.close()
            return True
    workbook.close()
    return False


def excluir_produto(produto_id):
    workbook, worksheet = abrir_aba(ABA_PRODUTOS)
    for linha in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=linha, column=1).value == produto_id:
            worksheet.delete_rows(linha)
            workbook.save(ARQUIVO_EXCEL)
            workbook.close()
            return True
    workbook.close()
    return False


def listar_vendas():
    workbook, worksheet = abrir_aba(ABA_VENDAS)
    vendas = []
    for linha in worksheet.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue
        vendas.append(normalizar_registro(linha, CABECALHO_VENDAS))
    workbook.close()
    return vendas


def validar_itens_carrinho(carrinho):
    if not carrinho:
        return False, "O carrinho esta vazio.", []

    workbook = load_workbook(ARQUIVO_EXCEL)
    produtos = workbook[ABA_PRODUTOS]
    itens_processados = []

    try:
        for item in carrinho:
            produto_linha = None
            for linha in range(2, produtos.max_row + 1):
                if produtos.cell(row=linha, column=1).value == item["produto_id"]:
                    produto_linha = linha
                    break

            if produto_linha is None:
                return False, "Produto nao encontrado.", []

            quantidade = int(item["quantidade"])
            if quantidade <= 0:
                return False, "Informe uma quantidade valida.", []

            nome_produto = str(produtos.cell(row=produto_linha, column=2).value)
            preco = float(produtos.cell(row=produto_linha, column=4).value)
            estoque_atual = int(float(produtos.cell(row=produto_linha, column=5).value))

            if estoque_atual < quantidade:
                return False, f"Estoque insuficiente para o produto {nome_produto}.", []

            itens_processados.append(
                {
                    "produto_id": item["produto_id"],
                    "produto_linha": produto_linha,
                    "produto_nome": nome_produto,
                    "quantidade": quantidade,
                    "valor_unitario": preco,
                    "valor_total": preco * quantidade,
                    "novo_estoque": estoque_atual - quantidade,
                }
            )
    finally:
        workbook.close()

    return True, "", itens_processados


def finalizar_compra(carrinho, cliente):
    ok, mensagem, itens_processados = validar_itens_carrinho(carrinho)
    if not ok:
        return False, mensagem, 0.0

    workbook = load_workbook(ARQUIVO_EXCEL)
    produtos = workbook[ABA_PRODUTOS]
    vendas = workbook[ABA_VENDAS]

    try:
        proximo_id_venda = proximo_id_generico(vendas)
        total_geral = 0.0
        data_venda = datetime.now().strftime("%d/%m/%Y %H:%M")

        for item in itens_processados:
            produtos.cell(row=item["produto_linha"], column=5, value=item["novo_estoque"])
            status_produto = "Disponivel" if item["novo_estoque"] > 0 else "Esgotado"
            produtos.cell(row=item["produto_linha"], column=6, value=status_produto)

            vendas.append(
                [
                    proximo_id_venda,
                    item["produto_id"],
                    item["produto_nome"],
                    cliente or "Consumidor final",
                    item["quantidade"],
                    f'{item["valor_unitario"]:.2f}',
                    f'{item["valor_total"]:.2f}',
                    "Finalizada",
                    data_venda,
                ]
            )
            proximo_id_venda += 1
            total_geral += item["valor_total"]

        workbook.save(ARQUIVO_EXCEL)
    finally:
        workbook.close()

    return True, f"Compra finalizada. Total: R$ {total_geral:.2f}", total_geral


def finalizar_venda(produto_id, cliente, quantidade):
    workbook = load_workbook(ARQUIVO_EXCEL)
    produtos = workbook[ABA_PRODUTOS]
    vendas = workbook[ABA_VENDAS]

    produto_linha = None
    for linha in range(2, produtos.max_row + 1):
        if produtos.cell(row=linha, column=1).value == produto_id:
            produto_linha = linha
            break

    if produto_linha is None:
        workbook.close()
        return False, "Produto nao encontrado.", 0.0

    nome_produto = str(produtos.cell(row=produto_linha, column=2).value)
    preco = float(produtos.cell(row=produto_linha, column=4).value)
    estoque_atual = int(float(produtos.cell(row=produto_linha, column=5).value))

    if quantidade <= 0:
        workbook.close()
        return False, "Informe uma quantidade valida.", 0.0

    if estoque_atual < quantidade:
        workbook.close()
        return False, "Estoque insuficiente para finalizar a venda.", 0.0

    novo_estoque = estoque_atual - quantidade
    produtos.cell(row=produto_linha, column=5, value=novo_estoque)
    if novo_estoque == 0:
        produtos.cell(row=produto_linha, column=6, value="Esgotado")

    novo_id = proximo_id_generico(vendas)
    valor_total = preco * quantidade
    vendas.append(
        [
            novo_id,
            produto_id,
            nome_produto,
            cliente or "Consumidor final",
            quantidade,
            f"{preco:.2f}",
            f"{valor_total:.2f}",
            "Finalizada",
            datetime.now().strftime("%d/%m/%Y %H:%M"),
        ]
    )

    workbook.save(ARQUIVO_EXCEL)
    workbook.close()
    return True, f"Venda finalizada. Total: R$ {valor_total:.2f}", valor_total


def criar_cliente(dados):
    workbook, worksheet = abrir_planilha()
    novo_id = proximo_id(worksheet)
    data_cadastro = datetime.now().strftime("%d/%m/%Y %H:%M")

    worksheet.append(
        [
            novo_id,
            dados["nome"],
            dados["cpf"],
            dados["email"],
            dados["telefone"],
            dados["idade"],
            dados["cidade"],
            dados["estado"].upper(),
            dados["status"] or STATUS_PADRAO,
            data_cadastro,
        ]
    )
    workbook.save(ARQUIVO_EXCEL)
    workbook.close()


def criar_cliente_rapido(nome, cpf):
    dados = {
        "nome": nome.strip(),
        "cpf": cpf.strip(),
        "email": "",
        "telefone": "",
        "idade": "",
        "cidade": "",
        "estado": "",
        "status": STATUS_PADRAO,
    }
    criar_cliente(dados)


def atualizar_cliente(cliente_id, dados):
    workbook, worksheet = abrir_planilha()

    for linha in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=linha, column=1).value == cliente_id:
            worksheet.cell(row=linha, column=2, value=dados["nome"])
            worksheet.cell(row=linha, column=3, value=dados["cpf"])
            worksheet.cell(row=linha, column=4, value=dados["email"])
            worksheet.cell(row=linha, column=5, value=dados["telefone"])
            worksheet.cell(row=linha, column=6, value=dados["idade"])
            worksheet.cell(row=linha, column=7, value=dados["cidade"])
            worksheet.cell(row=linha, column=8, value=dados["estado"].upper())
            worksheet.cell(row=linha, column=9, value=dados["status"] or STATUS_PADRAO)
            workbook.save(ARQUIVO_EXCEL)
            workbook.close()
            return True

    workbook.close()
    return False


def excluir_cliente(cliente_id):
    workbook, worksheet = abrir_planilha()

    for linha in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=linha, column=1).value == cliente_id:
            worksheet.delete_rows(linha)
            workbook.save(ARQUIVO_EXCEL)
            workbook.close()
            return True

    workbook.close()
    return False


def limpar_janela(raiz):
    for widget in raiz.winfo_children():
        widget.destroy()


class App:
    def __init__(self, raiz):
        self.raiz = raiz
        self.raiz.title("CRUD de Clientes com Excel")
        self.raiz.geometry("1360x720")
        self.raiz.minsize(1220, 680)
        self.raiz.configure(bg=FUNDO_JANELA)

        self.cliente_selecionado_id = None
        self.campos = {}
        self.busca_termo_var = tk.StringVar()
        self.carrinho = []
        self.logo_imagem = None
        self.logo_lateral_imagem = None
        self.logo_marca_dagua = None
        self.logo_fundo_lista = None
        self.fundo_animado = FundoAnimado(self.raiz, quantidade=4, margem=260)

        self.criar_estilos()
        self.criar_layout()
        self.carregar_tabela()

    def criar_estilos(self):
        estilo = ttk.Style()
        estilo.theme_use("clam")
        estilo.configure(
            "Treeview",
            rowheight=30,
            font=("Segoe UI", 10),
            background=FUNDO_TABELA,
            foreground=TEXTO_ESCURO,
            fieldbackground=FUNDO_TABELA,
            borderwidth=0,
        )
        estilo.map("Treeview", background=[("selected", ROXO_NEON)], foreground=[("selected", "#0b0214")])
        estilo.configure(
            "Treeview.Heading",
            font=("Segoe UI Semibold", 10),
            background=FUNDO_DESTAQUE,
            foreground=TEXTO_CLARO,
            relief="flat",
            borderwidth=0,
            padding=8,
        )
        estilo.map("Treeview.Heading", background=[("active", ROXO_NEON_FORTE)])
        estilo.configure(
            "TCombobox",
            fieldbackground=FUNDO_CAMPO,
            background=FUNDO_CAMPO,
            foreground=TEXTO_ESCURO,
            borderwidth=0,
            arrowsize=14,
        )

    def criar_layout(self):
        cabecalho = tk.Frame(self.raiz, bg=FUNDO_JANELA)
        cabecalho.pack(fill="x", padx=20, pady=(18, 10))

        faixa_marca = tk.Frame(cabecalho, bg=FUNDO_JANELA)
        faixa_marca.pack(side="left", anchor="w")

        self.criar_logo(faixa_marca)

        bloco_titulo = tk.Frame(faixa_marca, bg=FUNDO_JANELA)
        bloco_titulo.pack(side="left", anchor="center")

        titulo = tk.Label(
            bloco_titulo,
            text="Painel de Clientes",
            font=("Georgia", 26, "bold"),
            bg=FUNDO_JANELA,
            fg=TEXTO_CLARO,
        )
        titulo.pack(anchor="w")

        descricao = tk.Label(
            bloco_titulo,
            text="Gerencie seu cadastro com Excel em uma interface mais limpa, profissional e pronta para uso.",
            font=("Segoe UI", 11),
            bg=FUNDO_JANELA,
            fg=ROXO_NEON_FORTE,
        )
        descricao.pack(anchor="w", pady=(4, 0))

        acoes = tk.Frame(bloco_titulo, bg=FUNDO_JANELA)
        acoes.pack(anchor="w", pady=(12, 0))

        tk.Button(
            acoes,
            text="Produtos",
            command=self.abrir_janela_produtos,
            bg=AZUL_ACAO,
            fg="#14051f",
            relief="flat",
            padx=12,
            pady=6,
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left")

        tk.Button(
            acoes,
            text="Carrinho",
            command=self.abrir_janela_carrinho,
            bg=VERDE_ACAO,
            fg="#14051f",
            relief="flat",
            padx=12,
            pady=6,
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left", padx=(8, 0))

        self.criar_badge_administrador(cabecalho)

        container = tk.Frame(self.raiz, bg=FUNDO_JANELA)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        painel_formulario = tk.Frame(
            container,
            bg=FUNDO_CARD,
            bd=0,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
        )
        painel_formulario.pack(side="left", fill="y", padx=(0, 14))

        painel_lista = tk.Frame(
            container,
            bg=FUNDO_CARD,
            bd=0,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
        )
        painel_lista.pack(side="right", fill="both", expand=True)

        self.criar_formulario(painel_formulario)
        self.criar_area_lista(painel_lista)

    def criar_logo(self, parent):
        if not ARQUIVO_LOGO.exists():
            return

        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            self.logo_imagem = ImageTk.PhotoImage(imagem)
            logo_label = tk.Label(parent, image=self.logo_imagem, bg=FUNDO_JANELA)
            logo_label.pack(side="left", padx=(0, 14))
        except Exception:
            logo_label = tk.Label(
                parent,
                text="LOGO",
                font=("Segoe UI", 12, "bold"),
                bg=FUNDO_JANELA,
                fg=ROXO_NEON_FORTE,
            )
            logo_label.pack(side="left", padx=(0, 14))

    def criar_badge_administrador(self, parent):
        badge = tk.Frame(
            parent,
            bg=FUNDO_CARD,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
        )
        badge.pack(side="right", anchor="ne", padx=(16, 0))

        avatar = tk.Canvas(
            badge,
            width=54,
            height=54,
            bg=FUNDO_CARD,
            highlightthickness=0,
            bd=0,
        )
        avatar.pack(side="left", padx=(12, 10), pady=10)
        avatar.create_oval(9, 4, 45, 40, fill=ROXO_NEON, outline=ROXO_NEON_FORTE, width=2)
        avatar.create_oval(19, 12, 35, 28, fill=FUNDO_CARD, outline=FUNDO_CARD)
        avatar.create_arc(11, 24, 43, 52, start=0, extent=180, style="pieslice", fill=FUNDO_CARD, outline=FUNDO_CARD)

        textos = tk.Frame(badge, bg=FUNDO_CARD)
        textos.pack(side="left", padx=(0, 14), pady=10)

        tk.Label(
            textos,
            text="Usuario logado",
            font=("Segoe UI", 9, "bold"),
            bg=FUNDO_CARD,
            fg=ROXO_NEON_FORTE,
        ).pack(anchor="w")

        tk.Label(
            textos,
            text="Executado pelo administrador",
            font=("Segoe UI", 10),
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
        ).pack(anchor="w", pady=(2, 0))

        tk.Label(
            textos,
            text=EMAIL_DONO,
            font=("Segoe UI", 9),
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
        ).pack(anchor="w", pady=(2, 0))

    def criar_formulario(self, parent):
        subtitulo = tk.Label(
            parent,
            text="Formulario de Cliente",
            font=("Segoe UI", 15, "bold"),
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
        )
        subtitulo.pack(anchor="w", padx=16, pady=(16, 10))

        apoio = tk.Label(
            parent,
            text="Preencha os dados e use a tabela para selecionar um registro existente.",
            font=("Segoe UI", 9),
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
        )
        apoio.pack(anchor="w", padx=16, pady=(0, 10))

        linha_topo = tk.Frame(parent, bg=ROXO_NEON, height=2)
        linha_topo.pack(fill="x", padx=16, pady=(0, 12))

        form = tk.Frame(parent, bg=FUNDO_CARD)
        form.pack(fill="x", padx=16)

        labels = {
            "nome": "Nome",
            "cpf": "CPF",
            "email": "Email",
            "telefone": "Telefone",
            "idade": "Idade",
            "cidade": "Cidade",
            "estado": "Estado",
            "status": "Status",
        }

        for campo in CAMPOS_FORMULARIO:
            linha = tk.Frame(form, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
            linha.pack(fill="x", pady=4)

            label = tk.Label(
                linha,
                text=labels[campo],
                width=12,
                anchor="w",
                bg=FUNDO_CARD,
                fg=TEXTO_CLARO,
                font=("Segoe UI", 10),
            )
            label.pack(side="left")

            if campo == "status":
                widget = ttk.Combobox(linha, values=["Ativo", "Inativo"], state="readonly")
                widget.set(STATUS_PADRAO)
            else:
                widget = tk.Entry(
                    linha,
                    width=28,
                    font=("Segoe UI", 10),
                    bg=FUNDO_CAMPO,
                    fg=TEXTO_ESCURO,
                    relief="flat",
                    insertbackground=TEXTO_ESCURO,
                )

            widget.pack(side="left", fill="x", expand=True)
            self.campos[campo] = widget

        separador_botoes = tk.Frame(parent, bg=ROXO_NEON, height=2)
        separador_botoes.pack(fill="x", padx=16, pady=(12, 0))

        botoes = tk.Frame(parent, bg=FUNDO_CARD)
        botoes.pack(fill="x", padx=16, pady=16)

        botoes_config = [
            ("Novo / Limpar", self.limpar_formulario, AZUL_SUAVE, "#102a43"),
            ("Salvar", self.salvar_cliente, VERDE_ACAO, "#08121f"),
            ("Atualizar", self.editar_cliente, AZUL_ACAO, "#08121f"),
            ("Excluir", self.remover_cliente, VERMELHO_ACAO, TEXTO_CLARO),
        ]

        for texto, comando, cor, texto_cor in botoes_config:
            botao = tk.Button(
                botoes,
                text=texto,
                command=comando,
                bg=cor,
                fg=texto_cor,
                activebackground=cor,
                relief="flat",
                padx=12,
                pady=8,
                font=("Segoe UI", 10, "bold"),
            )
            botao.pack(fill="x", pady=4)

        self.criar_logo_rodape_formulario(parent)

    def criar_area_lista(self, parent):
        topo = tk.Frame(parent, bg=FUNDO_CARD)
        topo.pack(fill="x", padx=16, pady=16)

        tk.Label(
            topo,
            text="Busca global",
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
            font=("Segoe UI", 10),
        ).pack(side="left")

        tk.Label(
            topo,
            text="Pesquise por nome, CPF, email, cidade, estado, status ou qualquer outro campo.",
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
            font=("Segoe UI", 9),
        ).pack(side="left", padx=(10, 14))

        tk.Entry(
            topo,
            textvariable=self.busca_termo_var,
            width=30,
            font=("Segoe UI", 10),
            bg=FUNDO_CAMPO,
            fg=TEXTO_ESCURO,
            relief="flat",
            insertbackground=TEXTO_ESCURO,
        ).pack(side="left", padx=8)

        tk.Button(
            topo,
            text="Buscar",
            command=self.buscar_por_termo,
            bg=AZUL_ACAO,
            fg="#14051f",
            relief="flat",
            padx=12,
            pady=6,
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left")

        tk.Button(
            topo,
            text="Mostrar Todos",
            command=self.mostrar_todos_clientes,
            bg=FUNDO_DESTAQUE,
            fg=TEXTO_CLARO,
            relief="flat",
            padx=12,
            pady=6,
            font=("Segoe UI", 10, "bold"),
        ).pack(side="left", padx=8)

        colunas = CABECALHO
        linha_lista = tk.Frame(parent, bg=ROXO_NEON, height=2)
        linha_lista.pack(fill="x", padx=16, pady=(0, 12))

        area_tabela = tk.Frame(parent, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
        area_tabela.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        fundo_lista = tk.Frame(area_tabela, bg=FUNDO_TABELA)
        fundo_lista.grid(row=0, column=0, sticky="nsew")
        self.criar_marca_dagua_tabela(fundo_lista)

        self.tabela = ttk.Treeview(area_tabela, columns=colunas, show="headings")

        larguras = {
            "id": 50,
            "nome": 150,
            "cpf": 120,
            "email": 190,
            "telefone": 120,
            "idade": 60,
            "cidade": 120,
            "estado": 60,
            "status": 80,
            "data_cadastro": 130,
        }

        for coluna in colunas:
            self.tabela.heading(coluna, text=coluna.upper())
            self.tabela.column(coluna, width=larguras[coluna], anchor="center")

        scroll_y = ttk.Scrollbar(area_tabela, orient="vertical", command=self.tabela.yview)
        scroll_x = ttk.Scrollbar(area_tabela, orient="horizontal", command=self.tabela.xview)
        self.tabela.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        self.tabela.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        area_tabela.grid_rowconfigure(0, weight=1)
        area_tabela.grid_columnconfigure(0, weight=1)
        self.tabela.bind("<<TreeviewSelect>>", self.ao_selecionar_cliente)

    def criar_marca_dagua_tabela(self, parent):
        if not ARQUIVO_LOGO.exists():
            return

        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            largura = 300
            proporcao = largura / imagem.width
            altura = int(imagem.height * proporcao)
            imagem = imagem.resize((largura, altura), Image.LANCZOS)

            pixels = imagem.getdata()
            pixels_suaves = []
            for r, g, b, a in pixels:
                pixels_suaves.append((r, g, b, min(a, 28)))
            imagem.putdata(pixels_suaves)

            self.logo_marca_dagua = ImageTk.PhotoImage(imagem)
            self.logo_fundo_lista = tk.Label(parent, image=self.logo_marca_dagua, bg=FUNDO_TABELA, bd=0)
            self.logo_fundo_lista.place(relx=0.5, rely=0.5, anchor="center")
        except Exception:
            return

    def criar_logo_rodape_formulario(self, parent):
        if not ARQUIVO_LOGO.exists():
            return

        moldura_logo = tk.Frame(parent, bg=FUNDO_CARD)
        moldura_logo.pack(fill="x", padx=16, pady=(8, 16))

        separador = tk.Frame(moldura_logo, bg=ROXO_NEON, height=2)
        separador.pack(fill="x", pady=(0, 12))

        try:
            largura_painel = 320
            largura_logo = largura_painel - 32
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            proporcao = largura_logo / imagem.width
            nova_altura = int(imagem.height * proporcao)
            imagem = imagem.resize((largura_logo, nova_altura), Image.LANCZOS)
            self.logo_lateral_imagem = ImageTk.PhotoImage(imagem)
            logo_label = tk.Label(moldura_logo, image=self.logo_lateral_imagem, bg=FUNDO_CARD)
            logo_label.pack(anchor="center")
        except Exception:
            logo_label = tk.Label(
                moldura_logo,
                text="LOGO",
                font=("Segoe UI", 11, "bold"),
                bg=FUNDO_CARD,
                fg=ROXO_NEON_FORTE,
            )
            logo_label.pack(anchor="center")

    def obter_dados_formulario(self):
        dados = {}
        for campo, widget in self.campos.items():
            dados[campo] = widget.get().strip()
        return dados

    def validar_dados(self, dados):
        if not dados["nome"]:
            return "Informe o nome."
        if not dados["cpf"]:
            return "Informe o CPF."
        if len(normalizar_cpf(dados["cpf"])) != 11:
            return "Informe um CPF com 11 digitos."
        if not dados["email"]:
            return "Informe o email."
        if not email_aceito(dados["email"]):
            return "Informe um email valido no formato nome@dominio.com."
        if not dados["idade"]:
            return "Informe a idade."
        if not dados["idade"].isdigit():
            return "Informe a idade usando apenas numeros."
        if dados["estado"] and len(dados["estado"].strip()) != 2:
            return "Informe a UF com 2 letras."
        return None

    def preencher_formulario(self, cliente):
        for campo, widget in self.campos.items():
            if campo == "status":
                widget.set(cliente[campo] or STATUS_PADRAO)
            else:
                widget.delete(0, tk.END)
                widget.insert(0, cliente[campo])

    def limpar_formulario(self):
        self.cliente_selecionado_id = None
        for campo, widget in self.campos.items():
            if campo == "status":
                widget.set(STATUS_PADRAO)
            else:
                widget.delete(0, tk.END)
        self.tabela.selection_remove(self.tabela.selection())

    def carregar_tabela(self, clientes=None):
        for item in self.tabela.get_children():
            self.tabela.delete(item)

        lista_clientes = clientes if clientes is not None else listar_clientes()
        for cliente in lista_clientes:
            self.tabela.insert("", tk.END, values=[cliente[coluna] for coluna in CABECALHO])

    def mostrar_todos_clientes(self):
        self.busca_termo_var.set("")
        self.carregar_tabela()

    def ao_selecionar_cliente(self, _evento):
        selecionado = self.tabela.selection()
        if not selecionado:
            return

        valores = self.tabela.item(selecionado[0], "values")
        cliente = {coluna: valores[indice] for indice, coluna in enumerate(CABECALHO)}
        self.cliente_selecionado_id = int(cliente["id"])
        self.preencher_formulario(cliente)

    def salvar_cliente(self):
        dados = self.obter_dados_formulario()
        erro = self.validar_dados(dados)

        if erro:
            messagebox.showwarning("Validacao", erro)
            return

        if cpf_ja_cadastrado(dados["cpf"]):
            messagebox.showwarning("Validacao", "Ja existe um cliente com este CPF.")
            return

        criar_cliente(dados)
        self.carregar_tabela()
        self.limpar_formulario()
        messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso.")

    def editar_cliente(self):
        if self.cliente_selecionado_id is None:
            messagebox.showwarning("Selecao", "Selecione um cliente na tabela.")
            return

        dados = self.obter_dados_formulario()
        erro = self.validar_dados(dados)

        if erro:
            messagebox.showwarning("Validacao", erro)
            return

        if cpf_ja_cadastrado(dados["cpf"], self.cliente_selecionado_id):
            messagebox.showwarning("Validacao", "Ja existe outro cliente com este CPF.")
            return

        if atualizar_cliente(self.cliente_selecionado_id, dados):
            self.carregar_tabela()
            self.limpar_formulario()
            messagebox.showinfo("Sucesso", "Cliente atualizado com sucesso.")
            return

        messagebox.showerror("Erro", "Cliente nao encontrado.")

    def remover_cliente(self):
        if self.cliente_selecionado_id is None:
            messagebox.showwarning("Selecao", "Selecione um cliente na tabela.")
            return

        confirmar = messagebox.askyesno("Confirmacao", "Deseja realmente excluir este cliente?")
        if not confirmar:
            return

        if excluir_cliente(self.cliente_selecionado_id):
            self.carregar_tabela()
            self.limpar_formulario()
            messagebox.showinfo("Sucesso", "Cliente excluido com sucesso.")
            return

        messagebox.showerror("Erro", "Cliente nao encontrado.")

    def buscar_por_termo(self):
        termo = self.busca_termo_var.get().strip()
        if not termo:
            self.carregar_tabela()
            return

        clientes = buscar_clientes_por_termo(termo)
        if not clientes:
            self.carregar_tabela([])
            messagebox.showinfo("Busca", "Nenhum cliente encontrado para este termo.")
            return

        self.carregar_tabela(clientes)
        if len(clientes) == 1:
            self.preencher_formulario(clientes[0])
            self.cliente_selecionado_id = int(clientes[0]["id"])

    def abrir_janela_produtos(self):
        JanelaProdutos(self.raiz, self)

    def abrir_janela_carrinho(self):
        JanelaCarrinho(self.raiz, self)


class JanelaProdutos:
    def __init__(self, master, app):
        self.app = app
        self.modo_cliente = hasattr(app, "cliente_logado")
        self.janela = tk.Toplevel(master)
        self.janela.title("Produtos")
        self.janela.geometry("980x560")
        self.janela.configure(bg=FUNDO_JANELA)
        self.produto_id = None
        self.campos = {}
        self.quantidade_carrinho_var = tk.StringVar(value="1")
        self.produto_detalhe_imagem = None
        self.produto_atual = None
        self.criar_layout()
        self.carregar_produtos()

    def criar_layout(self):
        container = tk.Frame(self.janela, bg=FUNDO_JANELA)
        container.pack(fill="both", expand=True, padx=20, pady=20)

        painel_form = tk.Frame(container, bg=FUNDO_CARD, highlightthickness=2, highlightbackground=ROXO_NEON)
        painel_form.pack(side="left", fill="y", padx=(0, 12))

        painel_lista = tk.Frame(container, bg=FUNDO_CARD, highlightthickness=2, highlightbackground=ROXO_NEON)
        painel_lista.pack(side="right", fill="both", expand=True)

        titulo_form = "Catalogo de Produtos" if self.modo_cliente else "Cadastro de Produtos"
        tk.Label(painel_form, text=titulo_form, bg=FUNDO_CARD, fg=TEXTO_CLARO, font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=16, pady=(16, 12))

        if not self.modo_cliente:
            labels = [("nome", "Nome"), ("categoria", "Categoria"), ("preco", "Preco"), ("estoque", "Estoque"), ("status", "Status")]
            for campo, titulo in labels:
                linha = tk.Frame(painel_form, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
                linha.pack(fill="x", padx=16, pady=4)
                tk.Label(linha, text=titulo, width=12, anchor="w", bg=FUNDO_CARD, fg=TEXTO_CLARO).pack(side="left", padx=(8, 4), pady=8)
                if campo == "status":
                    widget = ttk.Combobox(linha, values=["Disponivel", "Esgotado", "Inativo"], state="readonly")
                    widget.set(STATUS_PRODUTO_PADRAO)
                else:
                    widget = tk.Entry(linha, bg=FUNDO_CAMPO, fg=TEXTO_ESCURO, relief="flat")
                widget.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=8)
                self.campos[campo] = widget

            botoes = tk.Frame(painel_form, bg=FUNDO_CARD)
            botoes.pack(fill="x", padx=16, pady=16)
            for texto, comando, cor in [
                ("Novo", self.limpar, AZUL_SUAVE),
                ("Salvar", self.salvar, VERDE_ACAO),
                ("Atualizar", self.atualizar, AZUL_ACAO),
                ("Excluir", self.excluir, VERMELHO_ACAO),
            ]:
                tk.Button(botoes, text=texto, command=comando, bg=cor, fg="#14051f" if cor != VERMELHO_ACAO else TEXTO_CLARO, relief="flat", font=("Segoe UI", 10, "bold")).pack(fill="x", pady=4)

            tk.Frame(painel_form, bg=ROXO_NEON, height=2).pack(fill="x", padx=16, pady=(8, 12))

        bloco_carrinho = tk.Frame(painel_form, bg=FUNDO_CARD)
        bloco_carrinho.pack(fill="x", padx=16, pady=(0, 16))

        tk.Label(
            bloco_carrinho,
            text="Compra do cliente",
            bg=FUNDO_CARD,
            fg=ROXO_NEON_FORTE,
            font=("Segoe UI", 12, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        linha_qtd = tk.Frame(bloco_carrinho, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
        linha_qtd.pack(fill="x", pady=(0, 10))
        tk.Label(linha_qtd, text="Quantidade", width=12, anchor="w", bg=FUNDO_CARD, fg=TEXTO_CLARO).pack(side="left", padx=(8, 4), pady=8)
        tk.Entry(linha_qtd, textvariable=self.quantidade_carrinho_var, bg=FUNDO_CAMPO, fg=TEXTO_ESCURO, relief="flat").pack(side="left", fill="x", expand=True, padx=(0, 8), pady=8)

        tk.Button(
            bloco_carrinho,
            text="Adicionar ao carrinho",
            command=self.adicionar_ao_carrinho,
            bg=AZUL_ACAO,
            fg="#14051f",
            relief="flat",
            font=("Segoe UI", 10, "bold"),
        ).pack(fill="x", pady=(0, 8))

        tk.Button(
            bloco_carrinho,
            text="Ir para o carrinho",
            command=self.ir_para_carrinho,
            bg=VERDE_ACAO,
            fg="#14051f",
            relief="flat",
            font=("Segoe UI", 10, "bold"),
        ).pack(fill="x")

        tk.Frame(painel_form, bg=ROXO_NEON, height=2).pack(fill="x", padx=16, pady=(14, 12))

        self.bloco_detalhes = tk.Frame(painel_form, bg=FUNDO_CARD)
        self.bloco_detalhes.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        tk.Label(
            self.bloco_detalhes,
            text="Detalhes do produto",
            bg=FUNDO_CARD,
            fg=ROXO_NEON_FORTE,
            font=("Segoe UI", 12, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        self.label_imagem_produto = tk.Label(
            self.bloco_detalhes,
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
            text="Selecione um produto para ver a imagem",
            compound="top",
            justify="center",
        )
        self.label_imagem_produto.pack(fill="x", pady=(0, 10))

        self.label_nome_detalhe = tk.Label(
            self.bloco_detalhes,
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
            font=("Segoe UI", 11, "bold"),
            text="Nome: -",
            anchor="w",
        )
        self.label_nome_detalhe.pack(fill="x", pady=2)

        self.label_preco_detalhe = tk.Label(
            self.bloco_detalhes,
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
            font=("Segoe UI", 10),
            text="Preco: -",
            anchor="w",
        )
        self.label_preco_detalhe.pack(fill="x", pady=2)

        self.label_categoria_detalhe = tk.Label(
            self.bloco_detalhes,
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
            font=("Segoe UI", 10),
            text="Categoria: -",
            anchor="w",
        )
        self.label_categoria_detalhe.pack(fill="x", pady=2)

        self.tabela_produtos = ttk.Treeview(painel_lista, columns=CABECALHO_PRODUTOS, show="headings")
        for coluna in CABECALHO_PRODUTOS:
            self.tabela_produtos.heading(coluna, text=coluna.upper())
            self.tabela_produtos.column(coluna, width=120, anchor="center")
        self.tabela_produtos.pack(fill="both", expand=True, padx=16, pady=16)
        self.tabela_produtos.bind("<<TreeviewSelect>>", self.selecionar)

    def obter_dados(self):
        return {campo: widget.get().strip() for campo, widget in self.campos.items()}

    def validar(self, dados):
        if not dados["nome"]:
            return "Informe o nome do produto."
        if not dados["preco"]:
            return "Informe o preco."
        if not dados["estoque"]:
            return "Informe o estoque."
        try:
            float(dados["preco"])
            int(dados["estoque"])
        except ValueError:
            return "Preco ou estoque invalidos."
        return None

    def carregar_produtos(self):
        for item in self.tabela_produtos.get_children():
            self.tabela_produtos.delete(item)
        for produto in listar_produtos():
            self.tabela_produtos.insert("", tk.END, values=[produto[col] for col in CABECALHO_PRODUTOS])

    def limpar(self):
        self.produto_id = None
        for campo, widget in self.campos.items():
            if campo == "status":
                widget.set(STATUS_PRODUTO_PADRAO)
            else:
                widget.delete(0, tk.END)

    def salvar(self):
        dados = self.obter_dados()
        erro = self.validar(dados)
        if erro:
            messagebox.showwarning("Produto", erro, parent=self.janela)
            return
        criar_produto(dados)
        self.carregar_produtos()
        self.limpar()
        messagebox.showinfo("Produto", "Produto cadastrado com sucesso.", parent=self.janela)

    def selecionar(self, _evento):
        selecionado = self.tabela_produtos.selection()
        if not selecionado:
            return
        valores = self.tabela_produtos.item(selecionado[0], "values")
        produto = {coluna: valores[indice] for indice, coluna in enumerate(CABECALHO_PRODUTOS)}
        self.produto_id = int(produto["id"])
        self.produto_atual = produto
        for campo, widget in self.campos.items():
            if campo == "status":
                widget.set(produto[campo])
            else:
                widget.delete(0, tk.END)
                widget.insert(0, produto[campo])
        self.atualizar_detalhes_produto(produto)

    def atualizar(self):
        if self.produto_id is None:
            messagebox.showwarning("Produto", "Selecione um produto.", parent=self.janela)
            return
        dados = self.obter_dados()
        erro = self.validar(dados)
        if erro:
            messagebox.showwarning("Produto", erro, parent=self.janela)
            return
        atualizar_produto(self.produto_id, dados)
        self.carregar_produtos()
        self.limpar()
        messagebox.showinfo("Produto", "Produto atualizado com sucesso.", parent=self.janela)

    def excluir(self):
        if self.produto_id is None:
            messagebox.showwarning("Produto", "Selecione um produto.", parent=self.janela)
            return
        excluir_produto(self.produto_id)
        self.carregar_produtos()
        self.limpar()
        messagebox.showinfo("Produto", "Produto excluido com sucesso.", parent=self.janela)

    def adicionar_ao_carrinho(self):
        if self.produto_id is None:
            messagebox.showwarning("Carrinho", "Selecione um produto na lista.", parent=self.janela)
            return
        try:
            quantidade = int(self.quantidade_carrinho_var.get().strip())
        except ValueError:
            messagebox.showwarning("Carrinho", "Quantidade invalida.", parent=self.janela)
            return
        if quantidade <= 0:
            messagebox.showwarning("Carrinho", "A quantidade deve ser maior que zero.", parent=self.janela)
            return

        produtos = listar_produtos()
        produto = next((item for item in produtos if int(item["id"]) == self.produto_id), None)
        if produto is None:
            messagebox.showwarning("Carrinho", "Produto nao encontrado.", parent=self.janela)
            return

        estoque = int(float(produto["estoque"]))
        if quantidade > estoque:
            messagebox.showwarning("Carrinho", "Quantidade maior que o estoque disponivel.", parent=self.janela)
            return

        existente = next((item for item in self.app.carrinho if item["produto_id"] == self.produto_id), None)
        if existente:
            if existente["quantidade"] + quantidade > estoque:
                messagebox.showwarning("Carrinho", "Quantidade total no carrinho excede o estoque.", parent=self.janela)
                return
            existente["quantidade"] += quantidade
        else:
            self.app.carrinho.append(
                {
                    "produto_id": self.produto_id,
                    "produto_nome": produto["nome"],
                    "valor_unitario": float(produto["preco"]),
                    "quantidade": quantidade,
                }
            )

        messagebox.showinfo("Carrinho", "Produto adicionado ao carrinho.", parent=self.janela)

    def ir_para_carrinho(self):
        JanelaCarrinho(self.janela, self.app)

    def atualizar_detalhes_produto(self, produto):
        self.label_nome_detalhe.configure(text=f'Nome: {produto["nome"]}')
        self.label_preco_detalhe.configure(text=f'Preco: R$ {produto["preco"]}')
        self.label_categoria_detalhe.configure(text=f'Categoria: {produto["categoria"]}')

        caminho_imagem = buscar_imagem_produto(produto["nome"])
        if caminho_imagem is None:
            self.produto_detalhe_imagem = None
            self.label_imagem_produto.configure(image="", text="Imagem nao encontrada para este produto")
            return

        try:
            imagem = Image.open(caminho_imagem).convert("RGBA")
            imagem.thumbnail((220, 180), Image.LANCZOS)
            self.produto_detalhe_imagem = ImageTk.PhotoImage(imagem)
            self.label_imagem_produto.configure(image=self.produto_detalhe_imagem, text="")
        except Exception:
            self.produto_detalhe_imagem = None
            self.label_imagem_produto.configure(image="", text="Nao foi possivel carregar a imagem")


class JanelaCarrinho:
    def __init__(self, master, app):
        self.app = app
        self.janela = tk.Toplevel(master)
        self.janela.title("Carrinho")
        self.janela.geometry("980x560")
        self.janela.configure(bg=FUNDO_JANELA)
        nome_cliente = app.cliente_logado["nome"] if hasattr(app, "cliente_logado") else ""
        self.cliente_var = tk.StringVar(value=nome_cliente)
        self.total_var = tk.StringVar(value="Total do carrinho: R$ 0.00")
        self.criar_layout()
        self.carregar_carrinho()

    def criar_layout(self):
        container = tk.Frame(self.janela, bg=FUNDO_JANELA)
        container.pack(fill="both", expand=True, padx=20, pady=20)

        painel_form = tk.Frame(container, bg=FUNDO_CARD, highlightthickness=2, highlightbackground=ROXO_NEON)
        painel_form.pack(side="left", fill="y", padx=(0, 12))

        painel_lista = tk.Frame(container, bg=FUNDO_CARD, highlightthickness=2, highlightbackground=ROXO_NEON)
        painel_lista.pack(side="right", fill="both", expand=True)

        tk.Label(painel_form, text="Carrinho do Cliente", bg=FUNDO_CARD, fg=TEXTO_CLARO, font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=16, pady=(16, 12))

        campos = tk.Frame(painel_form, bg=FUNDO_CARD)
        campos.pack(fill="x", padx=16)

        linha = tk.Frame(campos, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
        linha.pack(fill="x", pady=4)
        tk.Label(linha, text="Cliente", width=12, anchor="w", bg=FUNDO_CARD, fg=TEXTO_CLARO).pack(side="left", padx=(8, 4), pady=8)
        tk.Entry(linha, textvariable=self.cliente_var, bg=FUNDO_CAMPO, fg=TEXTO_ESCURO, relief="flat").pack(side="left", fill="x", expand=True, padx=(0, 8), pady=8)

        tk.Label(painel_form, textvariable=self.total_var, justify="left", wraplength=260, bg=FUNDO_CARD, fg=ROXO_NEON_FORTE, font=("Segoe UI", 12, "bold")).pack(fill="x", padx=16, pady=(12, 12))

        tk.Button(
            painel_form,
            text="Remover item selecionado",
            command=self.remover_item,
            bg=VERMELHO_ACAO,
            fg=TEXTO_CLARO,
            relief="flat",
            font=("Segoe UI", 10, "bold"),
        ).pack(fill="x", padx=16, pady=(0, 8))

        tk.Button(
            painel_form,
            text="Limpar carrinho",
            command=self.limpar_carrinho,
            bg=AZUL_SUAVE,
            fg="#14051f",
            relief="flat",
            font=("Segoe UI", 10, "bold"),
        ).pack(fill="x", padx=16, pady=(0, 8))

        tk.Button(
            painel_form,
            text="Finalizar Compra",
            command=self.finalizar,
            bg=VERDE_ACAO,
            fg="#14051f",
            relief="flat",
            font=("Segoe UI", 11, "bold"),
        ).pack(fill="x", padx=16, pady=(0, 16))

        colunas = ["produto_id", "produto_nome", "quantidade", "valor_unitario", "valor_total"]
        self.tabela_carrinho = ttk.Treeview(painel_lista, columns=colunas, show="headings")
        for coluna in colunas:
            self.tabela_carrinho.heading(coluna, text=coluna.upper())
            self.tabela_carrinho.column(coluna, width=140, anchor="center")
        self.tabela_carrinho.pack(fill="both", expand=True, padx=16, pady=16)

    def carregar_carrinho(self):
        for item in self.tabela_carrinho.get_children():
            self.tabela_carrinho.delete(item)

        total = 0.0
        for item in self.app.carrinho:
            valor_total = item["valor_unitario"] * item["quantidade"]
            total += valor_total
            self.tabela_carrinho.insert(
                "",
                tk.END,
                values=[
                    item["produto_id"],
                    item["produto_nome"],
                    item["quantidade"],
                    f'{item["valor_unitario"]:.2f}',
                    f"{valor_total:.2f}",
                ],
            )
        self.total_var.set(f"Total do carrinho: R$ {total:.2f}")

    def remover_item(self):
        selecionado = self.tabela_carrinho.selection()
        if not selecionado:
            messagebox.showwarning("Carrinho", "Selecione um item.", parent=self.janela)
            return
        valores = self.tabela_carrinho.item(selecionado[0], "values")
        produto_id = int(valores[0])
        self.app.carrinho = [item for item in self.app.carrinho if item["produto_id"] != produto_id]
        self.carregar_carrinho()

    def limpar_carrinho(self):
        self.app.carrinho.clear()
        self.carregar_carrinho()

    def finalizar(self):
        if not self.app.carrinho:
            messagebox.showwarning("Carrinho", "O carrinho esta vazio.", parent=self.janela)
            return

        ok, mensagem, total_geral = finalizar_compra(self.app.carrinho, self.cliente_var.get().strip())
        if not ok:
            messagebox.showwarning("Carrinho", mensagem, parent=self.janela)
            return

        self.app.carrinho.clear()
        self.carregar_carrinho()
        self.abrir_janela_pagamento(total_geral)

    def abrir_janela_pagamento(self, total_geral):
        JanelaPagamentoPix(self.janela, f"Compra finalizada com sucesso.\nTotal do carrinho: R$ {total_geral:.2f}")


class JanelaPagamentoPix:
    def __init__(self, master, mensagem_venda):
        self.janela = tk.Toplevel(master)
        self.janela.title("Pagamento PIX")
        self.janela.geometry("560x520")
        self.janela.configure(bg=FUNDO_JANELA)
        self.janela.resizable(False, False)
        self.logo_pix = None
        self.mensagem_venda = mensagem_venda
        self.criar_layout()

    def criar_layout(self):
        card = tk.Frame(
            self.janela,
            bg=FUNDO_CARD,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
        )
        card.place(relx=0.5, rely=0.5, anchor="center", width=470, height=430)

        topo = tk.Frame(card, bg=FUNDO_CARD)
        topo.pack(fill="x", padx=24, pady=(24, 8))

        self.criar_logo_pagamento(topo)

        textos = tk.Frame(topo, bg=FUNDO_CARD)
        textos.pack(side="left", padx=(12, 0))

        tk.Label(
            textos,
            text="Pagamento via PIX",
            font=("Georgia", 22, "bold"),
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
        ).pack(anchor="w")

        tk.Label(
            textos,
            text="Finalize o pagamento para concluir a compra.",
            font=("Segoe UI", 10),
            bg=FUNDO_CARD,
            fg=ROXO_NEON_FORTE,
        ).pack(anchor="w", pady=(4, 0))

        tk.Frame(card, bg=ROXO_NEON, height=2).pack(fill="x", padx=24, pady=(8, 18))

        tk.Label(
            card,
            text=self.mensagem_venda,
            justify="center",
            wraplength=380,
            font=("Segoe UI", 11, "bold"),
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
        ).pack(padx=24, pady=(0, 18))

        quadro_pix = tk.Frame(card, bg=FUNDO_JANELA, highlightthickness=1, highlightbackground=ROXO_NEON_FORTE)
        quadro_pix.pack(fill="x", padx=28, pady=(0, 18))

        tk.Label(
            quadro_pix,
            text="Chave PIX",
            font=("Segoe UI", 10, "bold"),
            bg=FUNDO_JANELA,
            fg=ROXO_NEON_FORTE,
        ).pack(anchor="w", padx=16, pady=(16, 6))

        entrada_pix = tk.Entry(
            quadro_pix,
            justify="center",
            font=("Segoe UI", 12, "bold"),
            bg=FUNDO_CAMPO,
            fg=TEXTO_ESCURO,
            relief="flat",
            insertbackground=TEXTO_ESCURO,
        )
        entrada_pix.pack(fill="x", padx=16, pady=(0, 14))
        entrada_pix.insert(0, CHAVE_PIX)
        entrada_pix.configure(state="readonly")

        tk.Label(
            card,
            text="Agradecemos pela preferencia.\nSua compra foi registrada com sucesso.",
            justify="center",
            font=("Segoe UI", 12),
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
        ).pack(padx=24, pady=(0, 22))

        tk.Button(
            card,
            text="Fechar",
            command=self.janela.destroy,
            bg=AZUL_ACAO,
            fg="#14051f",
            relief="flat",
            padx=16,
            pady=8,
            font=("Segoe UI", 11, "bold"),
        ).pack(pady=(0, 18))

    def criar_logo_pagamento(self, parent):
        if not ARQUIVO_LOGO.exists():
            return

        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            imagem = imagem.resize((96, 58), Image.LANCZOS)
            self.logo_pix = ImageTk.PhotoImage(imagem)
            tk.Label(parent, image=self.logo_pix, bg=FUNDO_CARD).pack(side="left")
        except Exception:
            pass


class LojaClienteApp:
    def __init__(self, raiz, cliente):
        self.raiz = raiz
        self.cliente_logado = cliente
        self.raiz.title("Area do Cliente")
        self.raiz.geometry("1180x640")
        self.raiz.minsize(1000, 580)
        self.raiz.configure(bg=FUNDO_JANELA)

        self.carrinho = []
        self.logo_imagem = None
        self.fundo_animado = FundoAnimado(self.raiz, quantidade=4, margem=260)
        self.criar_layout()

    def criar_layout(self):
        cabecalho = tk.Frame(self.raiz, bg=FUNDO_JANELA)
        cabecalho.pack(fill="x", padx=20, pady=(18, 10))

        faixa_marca = tk.Frame(cabecalho, bg=FUNDO_JANELA)
        faixa_marca.pack(side="left", anchor="w")
        self.criar_logo(faixa_marca)

        bloco_titulo = tk.Frame(faixa_marca, bg=FUNDO_JANELA)
        bloco_titulo.pack(side="left", anchor="center")
        tk.Label(bloco_titulo, text="Area do Cliente", font=("Georgia", 26, "bold"), bg=FUNDO_JANELA, fg=TEXTO_CLARO).pack(anchor="w")
        tk.Label(bloco_titulo, text=f'Bem-vindo, {self.cliente_logado["nome"]}. Escolha seus produtos e siga para o carrinho.', font=("Segoe UI", 11), bg=FUNDO_JANELA, fg=ROXO_NEON_FORTE).pack(anchor="w", pady=(4, 0))

        acoes = tk.Frame(bloco_titulo, bg=FUNDO_JANELA)
        acoes.pack(anchor="w", pady=(12, 0))
        tk.Button(acoes, text="Produtos", command=self.abrir_janela_produtos, bg=AZUL_ACAO, fg="#14051f", relief="flat", padx=12, pady=6, font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Button(acoes, text="Carrinho", command=self.abrir_janela_carrinho, bg=VERDE_ACAO, fg="#14051f", relief="flat", padx=12, pady=6, font=("Segoe UI", 10, "bold")).pack(side="left", padx=(8, 0))

        card = tk.Frame(self.raiz, bg=FUNDO_CARD, highlightthickness=2, highlightbackground=ROXO_NEON)
        card.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        tk.Label(card, text="Catalogo pronto para compra", bg=FUNDO_CARD, fg=TEXTO_CLARO, font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=24, pady=(24, 8))
        tk.Label(card, text="Use o botao Produtos para ver detalhes, adicionar itens e depois finalizar no Carrinho.", bg=FUNDO_CARD, fg=TEXTO_MUTED, font=("Segoe UI", 11)).pack(anchor="w", padx=24)

    def criar_logo(self, parent):
        if not ARQUIVO_LOGO.exists():
            return
        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            self.logo_imagem = ImageTk.PhotoImage(imagem)
            tk.Label(parent, image=self.logo_imagem, bg=FUNDO_JANELA).pack(side="left", padx=(0, 14))
        except Exception:
            pass

    def abrir_janela_produtos(self):
        JanelaProdutos(self.raiz, self)

    def abrir_janela_carrinho(self):
        JanelaCarrinho(self.raiz, self)


class EscolhaAcessoApp:
    def __init__(self, raiz):
        self.raiz = raiz
        self.raiz.title("Escolha de Acesso")
        self.raiz.geometry("720x460")
        self.raiz.minsize(680, 420)
        self.raiz.configure(bg=FUNDO_JANELA)

        self.logo_imagem = None
        self.fundo_animado = FundoAnimado(self.raiz, quantidade=4, margem=340)

        self.criar_layout()

    def criar_layout(self):
        container = tk.Frame(
            self.raiz,
            bg=FUNDO_CARD,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
            width=520,
            height=320,
        )
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        topo = tk.Frame(container, bg=FUNDO_CARD)
        topo.pack(fill="x", padx=24, pady=(24, 8))

        self.criar_logo_login(topo)

        titulo_bloco = tk.Frame(topo, bg=FUNDO_CARD)
        titulo_bloco.pack(side="left", anchor="center")

        tk.Label(
            titulo_bloco,
            text="Quem vai entrar?",
            font=("Georgia", 24, "bold"),
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
        ).pack(anchor="w")

        tk.Label(
            titulo_bloco,
            text="Escolha abaixo se o acesso sera de administrador ou de cliente.",
            font=("Segoe UI", 11),
            bg=FUNDO_CARD,
            fg=ROXO_NEON_FORTE,
        ).pack(anchor="w", pady=(4, 0))

        tk.Frame(container, bg=ROXO_NEON, height=2).pack(fill="x", padx=24, pady=(8, 20))

        botoes = tk.Frame(container, bg=FUNDO_CARD)
        botoes.pack(fill="x", padx=24, pady=(18, 24))

        tk.Button(
            botoes,
            text="Voce e administrador? Clique aqui",
            command=self.abrir_login_admin,
            bg=AZUL_ACAO,
            fg="#14051f",
            relief="flat",
            padx=14,
            pady=10,
            font=("Segoe UI", 11, "bold"),
        ).pack(fill="x", pady=(0, 10))

        tk.Button(
            botoes,
            text="Se for cliente clique aqui",
            command=self.abrir_login_cliente,
            bg=VERDE_ACAO,
            fg="#14051f",
            relief="flat",
            padx=14,
            pady=10,
            font=("Segoe UI", 11, "bold"),
        ).pack(fill="x", pady=(0, 10))

        tk.Button(
            botoes,
            text="Novo cliente",
            command=self.abrir_cadastro_cliente,
            bg=FUNDO_DESTAQUE,
            fg=TEXTO_CLARO,
            relief="flat",
            padx=14,
            pady=14,
            font=("Segoe UI", 11, "bold"),
        ).pack(fill="x")

        tk.Label(
            container,
            text="Acesso separado por perfil",
            font=("Segoe UI", 10, "bold"),
            bg=FUNDO_CARD,
            fg=TEXTO_MUTED,
        ).pack(pady=(4, 16))

    def criar_logo_login(self, parent):
        if not ARQUIVO_LOGO.exists():
            return

        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            imagem = imagem.resize((120, 72), Image.LANCZOS)
            self.logo_imagem = ImageTk.PhotoImage(imagem)
            tk.Label(parent, image=self.logo_imagem, bg=FUNDO_CARD).pack(side="left", padx=(0, 14))
        except Exception:
            pass

    def abrir_login_admin(self):
        limpar_janela(self.raiz)
        LoginAdministradorApp(self.raiz)

    def abrir_login_cliente(self):
        limpar_janela(self.raiz)
        LoginClienteApp(self.raiz)

    def abrir_cadastro_cliente(self):
        limpar_janela(self.raiz)
        CadastroClienteApp(self.raiz)


class LoginAdministradorApp:
    def __init__(self, raiz):
        self.raiz = raiz
        self.raiz.title("Login do Administrador")
        self.raiz.configure(bg=FUNDO_JANELA)

        self.logo_imagem = None
        self.email_var = tk.StringVar()
        self.senha_var = tk.StringVar()
        self.fundo_animado = FundoAnimado(self.raiz, quantidade=4, margem=340)
        self.criar_layout()

    def criar_layout(self):
        container = tk.Frame(
            self.raiz,
            bg=FUNDO_CARD,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
            width=520,
            height=320,
        )
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        topo = tk.Frame(container, bg=FUNDO_CARD)
        topo.pack(fill="x", padx=24, pady=(24, 8))
        self.criar_logo_login(topo)

        titulo_bloco = tk.Frame(topo, bg=FUNDO_CARD)
        titulo_bloco.pack(side="left", anchor="center")

        tk.Label(titulo_bloco, text="Acesso do Dono", font=("Georgia", 24, "bold"), bg=FUNDO_CARD, fg=TEXTO_CLARO).pack(anchor="w")
        tk.Label(titulo_bloco, text="Entre com email e senha para abrir o painel completo.", font=("Segoe UI", 11), bg=FUNDO_CARD, fg=ROXO_NEON_FORTE).pack(anchor="w", pady=(4, 0))

        tk.Frame(container, bg=ROXO_NEON, height=2).pack(fill="x", padx=24, pady=(8, 20))

        formulario = tk.Frame(container, bg=FUNDO_CARD)
        formulario.pack(fill="x", padx=24, pady=(0, 12))
        self.criar_campo_login(formulario, "Email", self.email_var, False)
        self.criar_campo_login(formulario, "Senha", self.senha_var, True)

        botoes = tk.Frame(container, bg=FUNDO_CARD)
        botoes.pack(fill="x", padx=24, pady=(8, 24))
        tk.Button(botoes, text="Entrar", command=self.validar_login, bg=AZUL_ACAO, fg="#14051f", relief="flat", padx=14, pady=10, font=("Segoe UI", 11, "bold")).pack(fill="x", pady=(0, 8))
        tk.Button(botoes, text="Voltar", command=self.voltar, bg=AZUL_SUAVE, fg="#14051f", relief="flat", padx=14, pady=8, font=("Segoe UI", 10, "bold")).pack(fill="x")

        self.raiz.bind("<Return>", lambda _event: self.validar_login())

    def criar_logo_login(self, parent):
        if not ARQUIVO_LOGO.exists():
            return
        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            imagem = imagem.resize((120, 72), Image.LANCZOS)
            self.logo_imagem = ImageTk.PhotoImage(imagem)
            tk.Label(parent, image=self.logo_imagem, bg=FUNDO_CARD).pack(side="left", padx=(0, 14))
        except Exception:
            pass

    def criar_campo_login(self, parent, titulo, variavel, ocultar):
        bloco = tk.Frame(parent, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
        bloco.pack(fill="x", pady=8)

        tk.Label(
            bloco,
            text=titulo,
            width=12,
            anchor="w",
            bg=FUNDO_CARD,
            fg=TEXTO_CLARO,
            font=("Segoe UI", 10),
        ).pack(side="left", padx=(10, 4), pady=10)

        entrada = tk.Entry(
            bloco,
            textvariable=variavel,
            show="*" if ocultar else "",
            font=("Segoe UI", 11),
            bg=FUNDO_CAMPO,
            fg=TEXTO_ESCURO,
            relief="flat",
            insertbackground=TEXTO_ESCURO,
        )
        entrada.pack(side="left", fill="x", expand=True, padx=(0, 10), pady=10)

    def validar_login(self):
        email = self.email_var.get().strip()
        senha = self.senha_var.get().strip()

        if email == EMAIL_DONO and senha == SENHA_DONO:
            limpar_janela(self.raiz)
            self.raiz.unbind("<Return>")
            App(self.raiz)
            return

        messagebox.showerror("Login invalido", "Email ou senha incorretos.")

    def voltar(self):
        self.raiz.unbind("<Return>")
        limpar_janela(self.raiz)
        EscolhaAcessoApp(self.raiz)


class LoginClienteApp:
    def __init__(self, raiz):
        self.raiz = raiz
        self.raiz.title("Login do Cliente")
        self.raiz.configure(bg=FUNDO_JANELA)
        self.logo_imagem = None
        self.nome_var = tk.StringVar()
        self.cpf_var = tk.StringVar()
        self.fundo_animado = FundoAnimado(self.raiz, quantidade=4, margem=340)
        self.criar_layout()

    def criar_layout(self):
        container = tk.Frame(
            self.raiz,
            bg=FUNDO_CARD,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
            width=520,
            height=320,
        )
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        topo = tk.Frame(container, bg=FUNDO_CARD)
        topo.pack(fill="x", padx=24, pady=(24, 8))
        self.criar_logo_login(topo)

        titulo_bloco = tk.Frame(topo, bg=FUNDO_CARD)
        titulo_bloco.pack(side="left", anchor="center")
        tk.Label(titulo_bloco, text="Acesso do Cliente", font=("Georgia", 24, "bold"), bg=FUNDO_CARD, fg=TEXTO_CLARO).pack(anchor="w")
        tk.Label(titulo_bloco, text="Informe seu nome e CPF para entrar na area de compras.", font=("Segoe UI", 11), bg=FUNDO_CARD, fg=ROXO_NEON_FORTE).pack(anchor="w", pady=(4, 0))

        tk.Frame(container, bg=ROXO_NEON, height=2).pack(fill="x", padx=24, pady=(8, 20))

        formulario = tk.Frame(container, bg=FUNDO_CARD)
        formulario.pack(fill="x", padx=24, pady=(0, 12))
        self.criar_campo_login(formulario, "Nome", self.nome_var)
        self.criar_campo_login(formulario, "CPF", self.cpf_var)

        botoes = tk.Frame(container, bg=FUNDO_CARD)
        botoes.pack(fill="x", padx=24, pady=(8, 24))
        tk.Button(botoes, text="Entrar", command=self.validar_login, bg=VERDE_ACAO, fg="#14051f", relief="flat", padx=14, pady=10, font=("Segoe UI", 11, "bold")).pack(fill="x", pady=(0, 8))
        tk.Button(botoes, text="Voltar", command=self.voltar, bg=AZUL_SUAVE, fg="#14051f", relief="flat", padx=14, pady=8, font=("Segoe UI", 10, "bold")).pack(fill="x")

        self.raiz.bind("<Return>", lambda _event: self.validar_login())

    def criar_logo_login(self, parent):
        if not ARQUIVO_LOGO.exists():
            return
        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            imagem = imagem.resize((120, 72), Image.LANCZOS)
            self.logo_imagem = ImageTk.PhotoImage(imagem)
            tk.Label(parent, image=self.logo_imagem, bg=FUNDO_CARD).pack(side="left", padx=(0, 14))
        except Exception:
            pass

    def criar_campo_login(self, parent, titulo, variavel):
        bloco = tk.Frame(parent, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
        bloco.pack(fill="x", pady=8)

        tk.Label(bloco, text=titulo, width=12, anchor="w", bg=FUNDO_CARD, fg=TEXTO_CLARO, font=("Segoe UI", 10)).pack(side="left", padx=(10, 4), pady=10)
        tk.Entry(bloco, textvariable=variavel, font=("Segoe UI", 11), bg=FUNDO_CAMPO, fg=TEXTO_ESCURO, relief="flat", insertbackground=TEXTO_ESCURO).pack(side="left", fill="x", expand=True, padx=(0, 10), pady=10)

    def validar_login(self):
        cliente = buscar_cliente_por_nome_cpf(self.nome_var.get(), self.cpf_var.get())
        if cliente is None:
            messagebox.showerror("Login invalido", "Nome ou CPF nao encontrados. Use o mesmo CPF cadastrado, com ou sem pontuacao.", parent=self.raiz)
            return
        limpar_janela(self.raiz)
        self.raiz.unbind("<Return>")
        LojaClienteApp(self.raiz, cliente)

    def voltar(self):
        self.raiz.unbind("<Return>")
        limpar_janela(self.raiz)
        EscolhaAcessoApp(self.raiz)


class CadastroClienteApp:
    def __init__(self, raiz):
        self.raiz = raiz
        self.raiz.title("Cadastro de Novo Cliente")
        self.raiz.configure(bg=FUNDO_JANELA)
        self.logo_imagem = None
        self.nome_var = tk.StringVar()
        self.cpf_var = tk.StringVar()
        self.fundo_animado = FundoAnimado(self.raiz, quantidade=4, margem=340)
        self.criar_layout()

    def criar_layout(self):
        container = tk.Frame(
            self.raiz,
            bg=FUNDO_CARD,
            highlightthickness=2,
            highlightbackground=ROXO_NEON,
            highlightcolor=ROXO_NEON_FORTE,
            width=520,
            height=320,
        )
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)

        topo = tk.Frame(container, bg=FUNDO_CARD)
        topo.pack(fill="x", padx=24, pady=(24, 8))
        self.criar_logo_login(topo)

        titulo_bloco = tk.Frame(topo, bg=FUNDO_CARD)
        titulo_bloco.pack(side="left", anchor="center")
        tk.Label(titulo_bloco, text="Novo Cliente", font=("Georgia", 24, "bold"), bg=FUNDO_CARD, fg=TEXTO_CLARO).pack(anchor="w")
        tk.Label(titulo_bloco, text="Cadastre nome e CPF para entrar na area de compras.", font=("Segoe UI", 11), bg=FUNDO_CARD, fg=ROXO_NEON_FORTE).pack(anchor="w", pady=(4, 0))

        tk.Frame(container, bg=ROXO_NEON, height=2).pack(fill="x", padx=24, pady=(8, 20))

        formulario = tk.Frame(container, bg=FUNDO_CARD)
        formulario.pack(fill="x", padx=24, pady=(0, 12))
        self.criar_campo_login(formulario, "Nome", self.nome_var)
        self.criar_campo_login(formulario, "CPF", self.cpf_var)

        botoes = tk.Frame(container, bg=FUNDO_CARD)
        botoes.pack(fill="x", padx=24, pady=(8, 24))
        tk.Button(botoes, text="Cadastrar", command=self.cadastrar, bg=VERDE_ACAO, fg="#14051f", relief="flat", padx=14, pady=10, font=("Segoe UI", 11, "bold")).pack(fill="x", pady=(0, 8))
        tk.Button(botoes, text="Voltar", command=self.voltar, bg=AZUL_SUAVE, fg="#14051f", relief="flat", padx=14, pady=8, font=("Segoe UI", 10, "bold")).pack(fill="x")

        self.raiz.bind("<Return>", lambda _event: self.cadastrar())

    def criar_logo_login(self, parent):
        if not ARQUIVO_LOGO.exists():
            return
        try:
            imagem = Image.open(ARQUIVO_LOGO).convert("RGBA")
            imagem = imagem.resize((120, 72), Image.LANCZOS)
            self.logo_imagem = ImageTk.PhotoImage(imagem)
            tk.Label(parent, image=self.logo_imagem, bg=FUNDO_CARD).pack(side="left", padx=(0, 14))
        except Exception:
            pass

    def criar_campo_login(self, parent, titulo, variavel):
        bloco = tk.Frame(parent, bg=FUNDO_CARD, highlightthickness=1, highlightbackground=ROXO_NEON)
        bloco.pack(fill="x", pady=8)
        tk.Label(bloco, text=titulo, width=12, anchor="w", bg=FUNDO_CARD, fg=TEXTO_CLARO, font=("Segoe UI", 10)).pack(side="left", padx=(10, 4), pady=10)
        tk.Entry(bloco, textvariable=variavel, font=("Segoe UI", 11), bg=FUNDO_CAMPO, fg=TEXTO_ESCURO, relief="flat", insertbackground=TEXTO_ESCURO).pack(side="left", fill="x", expand=True, padx=(0, 10), pady=10)

    def cadastrar(self):
        nome = self.nome_var.get().strip()
        cpf = self.cpf_var.get().strip()
        if not nome:
            messagebox.showwarning("Cadastro", "Informe o nome.", parent=self.raiz)
            return
        if not cpf:
            messagebox.showwarning("Cadastro", "Informe o CPF.", parent=self.raiz)
            return
        if len(normalizar_cpf(cpf)) != 11:
            messagebox.showwarning("Cadastro", "Informe um CPF com 11 digitos.", parent=self.raiz)
            return
        if buscar_cliente_por_cpf(cpf) is not None:
            messagebox.showwarning("Cadastro", "Ja existe um cliente com este CPF.", parent=self.raiz)
            return

        criar_cliente_rapido(nome, cpf)
        messagebox.showinfo("Cadastro", "Cliente cadastrado com sucesso.", parent=self.raiz)
        self.raiz.unbind("<Return>")
        limpar_janela(self.raiz)
        LoginClienteApp(self.raiz)

    def voltar(self):
        self.raiz.unbind("<Return>")
        limpar_janela(self.raiz)
        EscolhaAcessoApp(self.raiz)


def main():
    inicializar_planilha()
    raiz = tk.Tk()
    EscolhaAcessoApp(raiz)
    raiz.mainloop()


if __name__ == "__main__":
    main()
