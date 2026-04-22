from __future__ import annotations

from datetime import datetime
from email.message import EmailMessage
from functools import wraps
from pathlib import Path
import hashlib
import hmac
import json
import os
import re
import smtplib
import sqlite3
import ssl
from urllib import error as urllib_error
from urllib import request as urllib_request

from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_from_directory, session, url_for
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("WEB_APP_DATA_DIR", BASE_DIR))
DB_PATH = Path(os.getenv("WEB_APP_DB_PATH", DATA_DIR / "web_app.db"))
EXCEL_PATH = BASE_DIR / "clientes.xlsx"
CONFIG_PATH = BASE_DIR / "config.json"
EXTENSOES_IMAGEM = [".png", ".jpg", ".jpeg", ".webp"]


def carregar_configuracoes():
    configuracoes = {
        "admin_email": "admin@orquideas.local",
        "admin_password": "1234",
        "pix_key": "pix@orquideas.local",
        "secret_key": "orquideas-web-dev",
        "pagbank_token": "",
        "pagbank_webhook_token": "",
        "pagbank_notification_url": "",
        "pagbank_api_base": "https://sandbox.api.pagseguro.com",
        "loja_base_url": "",
        "smtp_host": "",
        "smtp_port": "587",
        "smtp_user": "",
        "smtp_password": "",
        "smtp_sender": "",
        "smtp_use_tls": "true",
    }
    if CONFIG_PATH.exists():
        try:
            with CONFIG_PATH.open("r", encoding="utf-8") as arquivo:
                conteudo = json.load(arquivo)
            if isinstance(conteudo, dict):
                configuracoes.update({chave: str(valor) for chave, valor in conteudo.items() if valor})
        except (OSError, json.JSONDecodeError):
            pass
    configuracoes["admin_email"] = os.getenv("ORQ_ADMIN_EMAIL", configuracoes["admin_email"]).strip()
    configuracoes["admin_password"] = os.getenv("ORQ_ADMIN_PASSWORD", configuracoes["admin_password"])
    configuracoes["pix_key"] = os.getenv("ORQ_PIX_KEY", configuracoes["pix_key"]).strip()
    configuracoes["secret_key"] = os.getenv("ORQ_WEB_SECRET_KEY", configuracoes["secret_key"])
    configuracoes["pagbank_token"] = os.getenv("PAGBANK_TOKEN", configuracoes["pagbank_token"]).strip()
    configuracoes["pagbank_webhook_token"] = os.getenv(
        "PAGBANK_WEBHOOK_TOKEN", configuracoes["pagbank_webhook_token"]
    ).strip()
    configuracoes["pagbank_notification_url"] = os.getenv(
        "PAGBANK_NOTIFICATION_URL", configuracoes["pagbank_notification_url"]
    ).strip()
    configuracoes["pagbank_api_base"] = os.getenv("PAGBANK_API_BASE", configuracoes["pagbank_api_base"]).strip().rstrip(
        "/"
    )
    configuracoes["loja_base_url"] = os.getenv("APP_BASE_URL", configuracoes["loja_base_url"]).strip().rstrip("/")
    configuracoes["smtp_host"] = os.getenv("SMTP_HOST", configuracoes["smtp_host"]).strip()
    configuracoes["smtp_port"] = os.getenv("SMTP_PORT", configuracoes["smtp_port"]).strip()
    configuracoes["smtp_user"] = os.getenv("SMTP_USER", configuracoes["smtp_user"]).strip()
    configuracoes["smtp_password"] = os.getenv("SMTP_PASSWORD", configuracoes["smtp_password"])
    configuracoes["smtp_sender"] = os.getenv("SMTP_SENDER", configuracoes["smtp_sender"]).strip()
    configuracoes["smtp_use_tls"] = os.getenv("SMTP_USE_TLS", configuracoes["smtp_use_tls"]).strip().lower()
    return configuracoes


CONFIG = carregar_configuracoes()
app = Flask(__name__)
app.config["SECRET_KEY"] = CONFIG["secret_key"]
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.getenv("FLASK_ENV") == "production"
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)


def agora_texto():
    return datetime.now().strftime("%d/%m/%Y %H:%M")


def parse_bool(valor):
    return str(valor or "").strip().lower() in {"1", "true", "yes", "sim", "on"}


def valor_configuracao_ativa(valor):
    texto = str(valor or "").strip()
    if not texto:
        return ""

    texto_normalizado = texto.lower()
    marcadores_exemplo = (
        "seu-",
        "sua-",
        "example",
        "exemplo",
        "troque-esta-senha",
        "smtp.seuprovedor.com",
        "usuario",
        "senha",
    )
    dominios_exemplo = (
        "orquideas.local",
        "seu-endereco-publico",
    )
    if any(marcador in texto_normalizado for marcador in marcadores_exemplo):
        return ""
    if any(dominio in texto_normalizado for dominio in dominios_exemplo):
        return ""
    return texto


def parse_float(valor, padrao=0.0):
    try:
        return float(str(valor).replace(",", "."))
    except (TypeError, ValueError):
        return padrao


def normalizar_email(email):
    return str(email or "").strip().lower()


def url_publica_base():
    base_url = valor_configuracao_ativa(CONFIG["loja_base_url"])
    if base_url:
        return base_url
    return request.host_url.rstrip("/")


def pagbank_configurado():
    return bool(valor_configuracao_ativa(CONFIG["pagbank_token"]))


def obter_pagbank_api_base():
    return (CONFIG["pagbank_api_base"] or "https://sandbox.api.pagseguro.com").rstrip("/")


def validar_assinatura_webhook():
    token = valor_configuracao_ativa(CONFIG["pagbank_webhook_token"])
    if not token:
        return True

    assinatura_recebida = request.headers.get("x-authenticity-token", "").strip()
    corpo = request.get_data(cache=True, as_text=True)
    if not assinatura_recebida or not corpo:
        return False

    manifesto = f"{token}-{corpo}"
    hash_esperado = hashlib.sha256(manifesto.encode("utf-8")).hexdigest()
    return hmac.compare_digest(hash_esperado, assinatura_recebida)


def pagbank_headers(media_type="application/json"):
    token = valor_configuracao_ativa(CONFIG["pagbank_token"])
    if not token:
        raise RuntimeError("Token do PagBank nao configurado.")
    return {
        "Authorization": f"Bearer {token}",
        "Accept": media_type,
        "Content-Type": "application/json",
        "User-Agent": "CasaDasOrquideas-PagBank/1.0",
    }


def extrair_erro_pagbank(corpo_resposta, padrao):
    try:
        dados = json.loads(corpo_resposta)
    except json.JSONDecodeError:
        return padrao
    mensagens = []
    for item in dados.get("error_messages", []):
        descricao = str(item.get("description") or item.get("error") or "").strip()
        if descricao:
            mensagens.append(descricao)
    if mensagens:
        return "; ".join(mensagens)
    return padrao


def chamar_api_pagbank(metodo, caminho_ou_url, payload=None, accept="application/json"):
    url = caminho_ou_url if str(caminho_ou_url).startswith("http") else f"{obter_pagbank_api_base()}{caminho_ou_url}"
    corpo = None if payload is None else json.dumps(payload, ensure_ascii=True).encode("utf-8")
    headers = pagbank_headers(accept)
    requisicao = urllib_request.Request(url, data=corpo, method=metodo.upper(), headers=headers)
    try:
        with urllib_request.urlopen(requisicao, timeout=30) as resposta:
            conteudo = resposta.read().decode("utf-8")
    except urllib_error.HTTPError as erro:
        conteudo = erro.read().decode("utf-8", errors="replace")
        raise RuntimeError(extrair_erro_pagbank(conteudo, f"PagBank retornou HTTP {erro.code}.")) from erro
    except urllib_error.URLError as erro:
        raise RuntimeError(f"Falha de conexao com o PagBank: {erro.reason}") from erro

    if accept == "application/json":
        try:
            return json.loads(conteudo)
        except json.JSONDecodeError as erro:
            raise RuntimeError("PagBank retornou uma resposta invalida.") from erro
    return conteudo.strip()


def extrair_link(qr_code, rel):
    for item in qr_code.get("links", []):
        if str(item.get("rel") or "").upper() == rel.upper():
            return str(item.get("href") or "").strip()
    return ""


def primeiro_qr_code(dados):
    qr_codes = dados.get("qr_codes") or dados.get("qr_code") or []
    return qr_codes[0] if qr_codes else {}


def parse_pagbank_datetime(valor):
    texto = str(valor or "").strip()
    if not texto:
        return None
    for formato in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            return datetime.strptime(texto, formato).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            continue
    return None


def status_pagbank_para_pedido(status_pagbank):
    return "pago" if str(status_pagbank or "").upper() == "PAID" else "pendente"


def telefone_pagbank(telefone):
    digitos = re.sub(r"\D", "", str(telefone or ""))
    if len(digitos) < 10:
        return None
    if len(digitos) in {10, 11}:
        return {
            "country": "55",
            "area": digitos[:2],
            "number": digitos[2:],
            "type": "MOBILE",
        }
    if len(digitos) == 12:
        return {
            "country": digitos[:-10] or "55",
            "area": digitos[-10:-8],
            "number": digitos[-8:],
            "type": "MOBILE",
        }
    if len(digitos) >= 13:
        return {
            "country": digitos[:-11] or "55",
            "area": digitos[-11:-9],
            "number": digitos[-9:],
            "type": "MOBILE",
        }
    return None


def enviar_email_pagamento_aprovado(destinatario, pedido_id, total):
    destinatario = normalizar_email(destinatario)
    if not destinatario:
        return False, "Cliente sem email cadastrado."
    smtp_host = valor_configuracao_ativa(CONFIG["smtp_host"])
    smtp_sender = valor_configuracao_ativa(CONFIG["smtp_sender"])
    smtp_user = valor_configuracao_ativa(CONFIG["smtp_user"])
    if not smtp_host or not smtp_sender:
        return False, "SMTP nao configurado."

    mensagem = EmailMessage()
    mensagem["Subject"] = f"Pagamento aprovado do pedido #{pedido_id}"
    mensagem["From"] = smtp_sender
    mensagem["To"] = destinatario
    mensagem.set_content(
        "\n".join(
            [
                "Ola!",
                "",
                f"Recebemos a confirmacao do pagamento Pix do pedido #{pedido_id}.",
                f"Valor aprovado: R$ {total:.2f}",
                "",
                "Seu pedido agora esta com status pago.",
                "",
                "Equipe Casa das Orquideas",
            ]
        )
    )

    porta = int(CONFIG["smtp_port"] or "587")
    usar_tls = parse_bool(CONFIG["smtp_use_tls"])
    contexto_ssl = ssl.create_default_context()
    with smtplib.SMTP(smtp_host, porta, timeout=20) as servidor:
        servidor.ehlo()
        if usar_tls:
            servidor.starttls(context=contexto_ssl)
            servidor.ehlo()
        if smtp_user:
            servidor.login(smtp_user, CONFIG["smtp_password"])
        servidor.send_message(mensagem)
    return True, "Email enviado."


def preparar_armazenamento():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


def coluna_existe(db, tabela, coluna):
    colunas = db.execute(f"PRAGMA table_info({tabela})").fetchall()
    return any(item["name"] == coluna for item in colunas)


def garantir_colunas_pix(db):
    colunas_sales = {
        "paid_at": "ALTER TABLE sales ADD COLUMN paid_at TEXT",
        "updated_at": "ALTER TABLE sales ADD COLUMN updated_at TEXT",
    }
    for coluna, sql in colunas_sales.items():
        if not coluna_existe(db, "sales", coluna):
            db.execute(sql)

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS pix_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL UNIQUE,
            mercado_pago_payment_id TEXT,
            pagbank_order_id TEXT,
            pagbank_charge_id TEXT,
            pagbank_qr_code_id TEXT,
            external_reference TEXT,
            status TEXT NOT NULL,
            status_detail TEXT,
            qr_code TEXT,
            qr_code_base64 TEXT,
            ticket_url TEXT,
            notification_email_sent INTEGER NOT NULL DEFAULT 0,
            raw_response TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            approved_at TEXT
        );
        """
    )
    colunas_pix = {
        "pagbank_order_id": "ALTER TABLE pix_payments ADD COLUMN pagbank_order_id TEXT",
        "pagbank_charge_id": "ALTER TABLE pix_payments ADD COLUMN pagbank_charge_id TEXT",
        "pagbank_qr_code_id": "ALTER TABLE pix_payments ADD COLUMN pagbank_qr_code_id TEXT",
    }
    for coluna, sql in colunas_pix.items():
        if not coluna_existe(db, "pix_payments", coluna):
            db.execute(sql)


@app.teardown_appcontext
def close_db(_error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def normalizar_cpf(cpf):
    return re.sub(r"\D", "", str(cpf or ""))


def encontrar_imagem_produto(nome_produto):
    for extensao in EXTENSOES_IMAGEM:
        caminho = BASE_DIR / f"{nome_produto}{extensao}"
        if caminho.exists():
            return caminho.name
    return None


def garantir_imagens_produtos(db):
    produtos_sem_imagem = db.execute("SELECT id, name FROM products WHERE image_path IS NULL OR image_path = ''").fetchall()
    houve_atualizacao = False
    for produto in produtos_sem_imagem:
        imagem = encontrar_imagem_produto(produto["name"])
        if not imagem:
            continue
        db.execute("UPDATE products SET image_path = ? WHERE id = ?", (imagem, produto["id"]))
        houve_atualizacao = True
    if houve_atualizacao:
        db.commit()


def garantir_produto_catalogo(db):
    produto = db.execute("SELECT id FROM products WHERE LOWER(name) = LOWER(?)", ("Cattleya hibrida",)).fetchone()
    if produto:
        db.execute(
            """
            UPDATE products
            SET category = ?, price = ?, stock = ?, status = ?, image_path = ?
            WHERE id = ?
            """,
            ("Orquidea", 80.0, 6, "Disponivel", "cattleya.site.jpeg", produto["id"]),
        )
    else:
        db.execute(
            """
            INSERT INTO products (name, category, price, stock, status, image_path, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "Cattleya hibrida",
                "Orquidea",
                80.0,
                6,
                "Disponivel",
                "cattleya.site.jpeg",
                datetime.now().strftime("%d/%m/%Y %H:%M"),
            ),
        )
    db.commit()


def init_db():
    preparar_armazenamento()
    db = get_db()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            cpf TEXT NOT NULL UNIQUE,
            email TEXT,
            phone TEXT,
            city TEXT,
            state TEXT,
            password_hash TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT,
            price REAL NOT NULL,
            stock INTEGER NOT NULL,
            status TEXT NOT NULL,
            image_path TEXT,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            client_name TEXT NOT NULL,
            total REAL NOT NULL,
            status TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            product_name TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            unit_price REAL NOT NULL,
            total_price REAL NOT NULL
        );
        """
    )
    garantir_colunas_pix(db)
    db.commit()
    importar_do_excel_se_necessario(db)
    garantir_produto_catalogo(db)
    garantir_imagens_produtos(db)


def importar_do_excel_se_necessario(db):
    if not EXCEL_PATH.exists():
        return
    clientes = db.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    produtos = db.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    if clientes or produtos:
        return

    workbook = load_workbook(EXCEL_PATH, data_only=True)
    try:
        if "clientes" in workbook.sheetnames:
            worksheet = workbook["clientes"]
            for linha in worksheet.iter_rows(min_row=2, values_only=True):
                if not linha or linha[0] is None:
                    continue
                nome = str(linha[1] or "").strip()
                cpf = normalizar_cpf(linha[2])
                if not nome or not cpf:
                    continue
                db.execute(
                    """
                    INSERT OR IGNORE INTO clients (name, cpf, email, phone, city, state, password_hash, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        nome,
                        cpf,
                        str(linha[3] or "").strip(),
                        str(linha[4] or "").strip(),
                        str(linha[6] or "").strip(),
                        str(linha[7] or "").strip().upper(),
                        generate_password_hash(cpf),
                        str(linha[9] or datetime.now().strftime("%d/%m/%Y %H:%M")),
                    ),
                )
        if "produtos" in workbook.sheetnames:
            worksheet = workbook["produtos"]
            for linha in worksheet.iter_rows(min_row=2, values_only=True):
                if not linha or linha[0] is None:
                    continue
                nome = str(linha[1] or "").strip()
                if not nome:
                    continue
                db.execute(
                    """
                    INSERT INTO products (name, category, price, stock, status, image_path, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        nome,
                        str(linha[2] or "").strip(),
                        float(linha[3] or 0),
                        int(float(linha[4] or 0)),
                        str(linha[5] or "Disponivel").strip(),
                        encontrar_imagem_produto(nome),
                        str(linha[6] or datetime.now().strftime("%d/%m/%Y %H:%M")),
                    ),
                )
    finally:
        workbook.close()
    db.commit()


@app.before_request
def before_request():
    init_db()


@app.context_processor
def inject_globals():
    return {
        "admin_logado": session.get("admin_logged_in", False),
        "cliente_logado": session.get("client_name"),
    }


def admin_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if not session.get("admin_logged_in"):
            flash("Entre como administrador para acessar essa area.", "warning")
            return redirect(url_for("admin_login"))
        return view(**kwargs)

    return wrapped_view


def client_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if not session.get("client_id"):
            flash("Entre como cliente para continuar.", "warning")
            return redirect(url_for("cliente_login", next=request.full_path.rstrip("?")))
        return view(**kwargs)

    return wrapped_view


def carrinho_atual():
    cart = session.get("cart", [])
    if not isinstance(cart, list):
        session["cart"] = []
        session.modified = True
        return session["cart"]

    cart_normalizado = []
    houve_ajuste = False

    for item in cart:
        try:
            product_id = int(item.get("product_id"))
            nome = str(item.get("name", "")).strip()
            unit_price = float(item.get("unit_price", 0))
            quantidade = int(item.get("quantity", 0))
        except (AttributeError, TypeError, ValueError):
            houve_ajuste = True
            continue

        if product_id <= 0 or not nome or unit_price < 0 or quantidade <= 0:
            houve_ajuste = True
            continue

        cart_normalizado.append(
            {
                "product_id": product_id,
                "name": nome,
                "unit_price": unit_price,
                "quantity": quantidade,
            }
        )

        if item != cart_normalizado[-1]:
            houve_ajuste = True

    if houve_ajuste or len(cart_normalizado) != len(cart):
        session["cart"] = cart_normalizado
        session.modified = True

    return session["cart"]


def obter_destino_pos_login():
    destino = request.values.get("next", "").strip()
    if not destino or not destino.startswith("/"):
        return url_for("loja_cliente")
    if destino.startswith("//"):
        return url_for("loja_cliente")
    return destino


def obter_destino_pos_carrinho():
    destino = request.form.get("next", "").strip()
    if destino == "cart":
        return url_for("ver_carrinho")
    return url_for("loja_cliente")


def adicionar_item_ao_carrinho(product_id, quantidade):
    if quantidade <= 0:
        return "Quantidade invalida.", "error"

    produto = get_db().execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if not produto:
        return "Produto nao encontrado.", "error"
    if produto["stock"] < quantidade:
        return "Estoque insuficiente para esse produto.", "error"

    cart = carrinho_atual()
    existente = next((item for item in cart if item["product_id"] == product_id), None)
    if existente:
        if existente["quantity"] + quantidade > produto["stock"]:
            return "A quantidade total no carrinho excede o estoque.", "error"
        existente["quantity"] += quantidade
    else:
        cart.append(
            {
                "product_id": product_id,
                "name": produto["name"],
                "unit_price": float(produto["price"]),
                "quantity": quantidade,
            }
        )

    session.modified = True
    return "Produto adicionado ao carrinho.", "success"


def buscar_pedido_por_id(sale_id, client_id=None):
    db = get_db()
    sql = """
        SELECT
            s.id,
            s.client_id,
            s.client_name,
            s.total,
            s.status,
            s.created_at,
            s.updated_at,
            s.paid_at,
            c.email AS client_email,
            c.cpf AS client_cpf,
            c.phone AS client_phone
        FROM sales s
        LEFT JOIN clients c ON c.id = s.client_id
        WHERE s.id = ?
    """
    parametros = [sale_id]
    if client_id is not None:
        sql += " AND s.client_id = ?"
        parametros.append(client_id)
    return db.execute(sql, tuple(parametros)).fetchone()


def buscar_pagamento_pix_por_pedido(sale_id):
    return get_db().execute("SELECT * FROM pix_payments WHERE sale_id = ?", (sale_id,)).fetchone()


def garantir_pedido_simples(valor, client_id=None, client_name=None):
    total = round(parse_float(valor), 2)
    if total <= 0:
        raise ValueError("Informe um valor maior que zero.")

    db = get_db()
    cliente_nome = client_name or session.get("client_name") or "Pedido Pix"
    cursor = db.execute(
        """
        INSERT INTO sales (client_id, client_name, total, status, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (client_id, cliente_nome, total, "pendente", agora_texto(), agora_texto()),
    )
    db.commit()
    return cursor.lastrowid


def payload_pix_pagbank(pedido):
    tax_id = normalizar_cpf(pedido["client_cpf"])
    if len(tax_id) not in {11, 14}:
        raise RuntimeError("O cliente precisa ter CPF ou CNPJ cadastrado para gerar Pix no PagBank.")

    notification_url = valor_configuracao_ativa(CONFIG["pagbank_notification_url"]) or (
        f"{url_publica_base()}{url_for('webhook_pagbank')}"
    )
    email_pagador = normalizar_email(pedido["client_email"]) or CONFIG["admin_email"]
    payload = {
        "reference_id": str(pedido["id"]),
        "customer": {
            "name": str(pedido["client_name"] or "Cliente")[:120],
            "email": email_pagador,
            "tax_id": tax_id,
        },
        "items": [
            {
                "reference_id": f"pedido-{pedido['id']}",
                "name": f"Pedido #{pedido['id']} - Casa das Orquideas",
                "quantity": 1,
                "unit_amount": int(round(float(pedido["total"]) * 100)),
            }
        ],
        "qr_codes": [
            {
                "amount": {
                    "value": int(round(float(pedido["total"]) * 100)),
                },
                "arrangements": ["PAGBANK"],
            }
        ],
        "notification_urls": [notification_url],
    }
    telefone = telefone_pagbank(pedido["client_phone"])
    if telefone:
        payload["customer"]["phones"] = [telefone]
    return payload


def criar_pagamento_pix_pagbank(pedido):
    dados = chamar_api_pagbank("POST", "/orders", payload=payload_pix_pagbank(pedido))
    qr_code_info = primeiro_qr_code(dados)
    charge = (dados.get("charges") or [{}])[0]
    qr_code = str(qr_code_info.get("text") or "").strip()
    qr_code_base64 = ""
    link_base64 = extrair_link(qr_code_info, "QRCODE.BASE64")
    if link_base64:
        qr_code_base64 = chamar_api_pagbank("GET", link_base64, accept="text/plain")
    if not dados.get("id") or not qr_code:
        raise RuntimeError("PagBank nao retornou os dados completos do Pix.")

    agora = agora_texto()
    db = get_db()
    db.execute(
        """
        INSERT INTO pix_payments (
            sale_id,
            mercado_pago_payment_id,
            pagbank_order_id,
            pagbank_charge_id,
            pagbank_qr_code_id,
            external_reference,
            status,
            status_detail,
            qr_code,
            qr_code_base64,
            ticket_url,
            raw_response,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(sale_id) DO UPDATE SET
            mercado_pago_payment_id = excluded.mercado_pago_payment_id,
            pagbank_order_id = excluded.pagbank_order_id,
            pagbank_charge_id = excluded.pagbank_charge_id,
            pagbank_qr_code_id = excluded.pagbank_qr_code_id,
            external_reference = excluded.external_reference,
            status = excluded.status,
            status_detail = excluded.status_detail,
            qr_code = excluded.qr_code,
            qr_code_base64 = excluded.qr_code_base64,
            ticket_url = excluded.ticket_url,
            raw_response = excluded.raw_response,
            updated_at = excluded.updated_at
        """,
        (
            pedido["id"],
            str(charge.get("id") or dados["id"]),
            str(dados["id"]),
            str(charge.get("id") or ""),
            str(qr_code_info.get("id") or ""),
            str(dados.get("reference_id") or pedido["id"]),
            str(charge.get("status") or "WAITING").upper(),
            str(charge.get("status") or "WAITING"),
            qr_code,
            qr_code_base64,
            extrair_link(qr_code_info, "QRCODE.PNG"),
            json.dumps(dados, ensure_ascii=True),
            agora,
            agora,
        ),
    )
    db.execute(
        "UPDATE sales SET status = ?, updated_at = ? WHERE id = ?",
        ("pendente", agora, pedido["id"]),
    )
    db.commit()
    return dados


def atualizar_status_pagamento_pagbank(pedido_pagbank):
    db = get_db()
    order_id = str(pedido_pagbank.get("id") or "").strip()
    external_reference = str(pedido_pagbank.get("reference_id") or "").strip()
    if not external_reference:
        return None

    sale_id = int(external_reference)
    pedido = buscar_pedido_por_id(sale_id)
    if not pedido:
        return None

    qr_code_info = primeiro_qr_code(pedido_pagbank)
    charge = (pedido_pagbank.get("charges") or [{}])[0]
    charge_id = str(charge.get("id") or "").strip()
    status_pagbank = str(charge.get("status") or "WAITING").upper()
    status_pedido = status_pagbank_para_pedido(status_pagbank)
    agora = agora_texto()

    db.execute(
        """
        INSERT INTO pix_payments (
            sale_id,
            mercado_pago_payment_id,
            pagbank_order_id,
            pagbank_charge_id,
            pagbank_qr_code_id,
            external_reference,
            status,
            status_detail,
            qr_code,
            qr_code_base64,
            ticket_url,
            raw_response,
            created_at,
            updated_at,
            approved_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(sale_id) DO UPDATE SET
            mercado_pago_payment_id = excluded.mercado_pago_payment_id,
            pagbank_order_id = excluded.pagbank_order_id,
            pagbank_charge_id = excluded.pagbank_charge_id,
            pagbank_qr_code_id = excluded.pagbank_qr_code_id,
            external_reference = excluded.external_reference,
            status = excluded.status,
            status_detail = excluded.status_detail,
            qr_code = CASE WHEN excluded.qr_code != '' THEN excluded.qr_code ELSE pix_payments.qr_code END,
            qr_code_base64 = CASE
                WHEN excluded.qr_code_base64 != '' THEN excluded.qr_code_base64
                ELSE pix_payments.qr_code_base64
            END,
            ticket_url = CASE WHEN excluded.ticket_url != '' THEN excluded.ticket_url ELSE pix_payments.ticket_url END,
            raw_response = excluded.raw_response,
            updated_at = excluded.updated_at,
            approved_at = CASE
                WHEN excluded.approved_at IS NOT NULL THEN excluded.approved_at
                ELSE pix_payments.approved_at
            END
        """,
        (
            sale_id,
            charge_id or order_id,
            order_id,
            charge_id,
            str(qr_code_info.get("id") or ""),
            external_reference,
            status_pagbank,
            str(charge.get("status") or status_pagbank),
            str(qr_code_info.get("text") or ""),
            "",
            extrair_link(qr_code_info, "QRCODE.PNG"),
            json.dumps(pedido_pagbank, ensure_ascii=True),
            agora,
            agora,
            parse_pagbank_datetime(charge.get("paid_at")) or (agora if status_pagbank == "PAID" else None),
        ),
    )
    db.execute(
        "UPDATE sales SET status = ?, updated_at = ?, paid_at = CASE WHEN ? = 'pago' THEN ? ELSE paid_at END WHERE id = ?",
        (status_pedido, agora, status_pedido, agora, sale_id),
    )

    pagamento_local = buscar_pagamento_pix_por_pedido(sale_id)
    if (
        status_pedido == "pago"
        and pedido["client_email"]
        and pagamento_local
        and not int(pagamento_local["notification_email_sent"])
    ):
        try:
            enviado, _mensagem = enviar_email_pagamento_aprovado(
                pedido["client_email"], sale_id, float(pedido["total"])
            )
        except Exception:
            enviado = False
        if enviado:
            db.execute(
                "UPDATE pix_payments SET notification_email_sent = 1, updated_at = ? WHERE sale_id = ?",
                (agora_texto(), sale_id),
            )

    db.commit()
    return buscar_pedido_por_id(sale_id)


def consultar_pedido_pagbank(order_id):
    if not order_id:
        raise RuntimeError("Pedido PagBank nao informado.")
    return chamar_api_pagbank("GET", f"/orders/{order_id}")


@app.route("/health")
def healthcheck():
    db = get_db()
    db.execute("SELECT 1").fetchone()
    return {"status": "ok"}, 200


@app.route("/")
def index():
    db = get_db()
    contagens = {
        "clientes": db.execute("SELECT COUNT(*) FROM clients").fetchone()[0],
        "produtos": db.execute("SELECT COUNT(*) FROM products").fetchone()[0],
        "vendas": db.execute("SELECT COUNT(*) FROM sales").fetchone()[0],
    }
    destaques = db.execute(
        "SELECT id, name, category, price, stock, status, image_path FROM products ORDER BY id DESC LIMIT 6"
    ).fetchall()
    catalogo_publico = {
        "nome": "Cattleya hibrida",
        "preco": "80,00",
        "imagem": "cattleya.site.jpeg",
        "descricao": (
            "Orquidea de flores grandes, perfumadas e vibrantes, ideal para quem quer destacar "
            "a beleza natural do ambiente com um toque elegante."
        ),
        "ambiente": "Vai muito bem em locais bem iluminados, com luz indireta e boa ventilacao.",
        "cuidados": "Rega moderada e substrato leve, sempre deixando secar entre uma irrigacao e outra.",
    }
    produto_catalogo = db.execute(
        "SELECT id FROM products WHERE LOWER(name) = LOWER(?) LIMIT 1",
        (catalogo_publico["nome"],),
    ).fetchone()
    return render_template(
        "index.html",
        contagens=contagens,
        destaques=destaques,
        catalogo_publico=catalogo_publico,
        catalogo_produto_id=produto_catalogo["id"] if produto_catalogo else None,
    )


@app.route("/assets/<path:filename>")
def asset_file(filename):
    if Path(filename).suffix.lower() not in EXTENSOES_IMAGEM:
        return "", 404
    return send_from_directory(BASE_DIR, filename)


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        senha = request.form.get("password", "")
        if email == CONFIG["admin_email"] and senha == CONFIG["admin_password"]:
            session.clear()
            session["admin_logged_in"] = True
            flash("Login de administrador realizado.", "success")
            return redirect(url_for("admin_dashboard"))
        flash("Email ou senha invalidos.", "error")
    return render_template("admin_login.html")


@app.route("/cliente/login", methods=["GET", "POST"])
def cliente_login():
    proximo_destino = obter_destino_pos_login()
    if request.method == "POST":
        cpf = normalizar_cpf(request.form.get("cpf"))
        senha = request.form.get("password", "")
        cliente = get_db().execute("SELECT * FROM clients WHERE cpf = ?", (cpf,)).fetchone()
        if cliente and check_password_hash(cliente["password_hash"], senha):
            session.clear()
            session["client_id"] = cliente["id"]
            session["client_name"] = cliente["name"]
            session["cart"] = []
            flash("Login realizado com sucesso.", "success")
            return redirect(proximo_destino)
        flash("CPF ou senha invalidos.", "error")
    return render_template("cliente_login.html", proximo_destino=proximo_destino)


@app.route("/cliente/cadastro", methods=["GET", "POST"])
def cliente_cadastro():
    proximo_destino = obter_destino_pos_login()
    if request.method == "POST":
        nome = request.form.get("name", "").strip()
        cpf = normalizar_cpf(request.form.get("cpf"))
        email = request.form.get("email", "").strip()
        cidade = request.form.get("city", "").strip()
        estado = request.form.get("state", "").strip().upper()
        senha = request.form.get("password", "")
        confirmar = request.form.get("password_confirm", "")
        if not nome or len(cpf) != 11 or not senha:
            flash("Preencha nome, CPF valido e senha.", "error")
            return render_template("cliente_cadastro.html", proximo_destino=proximo_destino)
        if senha != confirmar:
            flash("As senhas nao conferem.", "error")
            return render_template("cliente_cadastro.html", proximo_destino=proximo_destino)
        db = get_db()
        try:
            db.execute(
                """
                INSERT INTO clients (name, cpf, email, phone, city, state, password_hash, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    nome,
                    cpf,
                    email,
                    request.form.get("phone", "").strip(),
                    cidade,
                    estado,
                    generate_password_hash(senha),
                    datetime.now().strftime("%d/%m/%Y %H:%M"),
                ),
            )
            db.commit()
        except sqlite3.IntegrityError:
            flash("Ja existe um cliente com esse CPF.", "error")
            return render_template("cliente_cadastro.html", proximo_destino=proximo_destino)
        flash("Cadastro concluido. Agora voce ja pode entrar.", "success")
        return redirect(url_for("cliente_login", next=proximo_destino))
    return render_template("cliente_cadastro.html", proximo_destino=proximo_destino)


@app.route("/logout")
def logout():
    session.clear()
    flash("Sessao encerrada.", "success")
    return redirect(url_for("index"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    db = get_db()
    clientes = db.execute("SELECT * FROM clients ORDER BY id DESC").fetchall()
    produtos = db.execute("SELECT * FROM products ORDER BY id DESC").fetchall()
    vendas = db.execute("SELECT * FROM sales ORDER BY id DESC LIMIT 20").fetchall()
    return render_template("admin_dashboard.html", clientes=clientes, produtos=produtos, vendas=vendas)


@app.post("/admin/produtos/novo")
@admin_required
def admin_novo_produto():
    nome = request.form.get("name", "").strip()
    categoria = request.form.get("category", "").strip()
    status = request.form.get("status", "Disponivel").strip()
    try:
        preco = float(request.form.get("price", "0"))
        estoque = int(request.form.get("stock", "0"))
    except ValueError:
        flash("Preco ou estoque invalidos.", "error")
        return redirect(url_for("admin_dashboard"))

    if not nome:
        flash("Informe o nome do produto.", "error")
        return redirect(url_for("admin_dashboard"))

    db = get_db()
    db.execute(
        """
        INSERT INTO products (name, category, price, stock, status, image_path, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            nome,
            categoria,
            preco,
            estoque,
            status,
            encontrar_imagem_produto(nome),
            datetime.now().strftime("%d/%m/%Y %H:%M"),
        ),
    )
    db.commit()
    flash("Produto criado no site.", "success")
    return redirect(url_for("admin_dashboard"))


@app.post("/admin/clientes/novo")
@admin_required
def admin_novo_cliente():
    nome = request.form.get("name", "").strip()
    cpf = normalizar_cpf(request.form.get("cpf"))
    senha = request.form.get("password", "").strip()
    if not nome or len(cpf) != 11:
        flash("Informe nome e CPF valido.", "error")
        return redirect(url_for("admin_dashboard"))

    db = get_db()
    try:
        db.execute(
            """
            INSERT INTO clients (name, cpf, email, phone, city, state, password_hash, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                nome,
                cpf,
                request.form.get("email", "").strip(),
                request.form.get("phone", "").strip(),
                request.form.get("city", "").strip(),
                request.form.get("state", "").strip().upper(),
                generate_password_hash(senha or cpf),
                datetime.now().strftime("%d/%m/%Y %H:%M"),
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        flash("CPF ja cadastrado no site.", "error")
        return redirect(url_for("admin_dashboard"))

    flash("Cliente criado no site.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/loja")
@client_required
def loja_cliente():
    produtos = get_db().execute(
        "SELECT * FROM products WHERE status != 'Inativo' ORDER BY name"
    ).fetchall()
    itens_carrinho = sum(item["quantity"] for item in carrinho_atual())
    return render_template("cliente_loja.html", produtos=produtos, itens_carrinho=itens_carrinho)


@app.post("/carrinho/adicionar/<int:product_id>")
@client_required
def adicionar_carrinho(product_id):
    try:
        quantidade = int(request.form.get("quantity", "1"))
    except ValueError:
        quantidade = 0
    mensagem, categoria = adicionar_item_ao_carrinho(product_id, quantidade)
    flash(mensagem, categoria)
    return redirect(obter_destino_pos_carrinho())


@app.post("/comprar-agora/<int:product_id>")
@client_required
def comprar_agora(product_id):
    try:
        quantidade = int(request.form.get("quantity", "1"))
    except ValueError:
        quantidade = 0
    mensagem, categoria = adicionar_item_ao_carrinho(product_id, quantidade)
    flash(mensagem, categoria)
    return redirect(url_for("ver_carrinho"))


@app.route("/carrinho")
@client_required
def ver_carrinho():
    cart = carrinho_atual()
    total = sum(item["unit_price"] * item["quantity"] for item in cart)
    return render_template(
        "carrinho.html",
        cart=cart,
        total=total,
        pix_key=CONFIG["pix_key"],
        itens_carrinho=sum(item["quantity"] for item in cart),
    )


@app.post("/criar_pix")
def criar_pix():
    dados = request.get_json(silent=True) or request.form
    pedido_id = str(dados.get("pedido_id", "")).strip()
    valor_informado = dados.get("valor")
    client_id = session.get("client_id")

    try:
        if pedido_id:
            pedido = buscar_pedido_por_id(int(pedido_id), client_id=client_id if client_id else None)
            if not pedido:
                return jsonify({"erro": "Pedido nao encontrado."}), 404
            if valor_informado is not None and round(parse_float(valor_informado), 2) != round(float(pedido["total"]), 2):
                return jsonify({"erro": "O valor informado difere do total do pedido."}), 400
        else:
            novo_pedido_id = garantir_pedido_simples(
                valor_informado,
                client_id=client_id,
                client_name=session.get("client_name"),
            )
            pedido = buscar_pedido_por_id(novo_pedido_id, client_id=client_id if client_id else None)
    except ValueError as erro:
        return jsonify({"erro": str(erro)}), 400

    pagamento_existente = buscar_pagamento_pix_por_pedido(int(pedido["id"]))
    if pagamento_existente and str(pagamento_existente["status"]).upper() == "PAID":
        return jsonify(
            {
                "pedido_id": int(pedido["id"]),
                "payment_id": pagamento_existente["pagbank_charge_id"] or pagamento_existente["pagbank_order_id"],
                "status": "PAID",
                "mensagem": "Pagamento ja aprovado.",
            }
        )

    if pagamento_existente and pagamento_existente["qr_code"] and str(pagamento_existente["status"]).upper() in {
        "WAITING",
        "AUTHORIZED",
    }:
        return jsonify(
            {
                "pedido_id": int(pedido["id"]),
                "payment_id": pagamento_existente["pagbank_charge_id"] or pagamento_existente["pagbank_order_id"],
                "status": pagamento_existente["status"],
                "qr_code": pagamento_existente["qr_code"],
                "qr_code_base64": pagamento_existente["qr_code_base64"],
                "mensagem": "Pagamento Pix ja gerado para este pedido.",
            }
        )

    try:
        pedido_pagbank = criar_pagamento_pix_pagbank(pedido)
    except RuntimeError as erro:
        return jsonify({"erro": str(erro)}), 500
    except Exception as erro:  # pragma: no cover - depende da API externa
        return jsonify({"erro": f"Falha ao gerar Pix no PagBank: {erro}"}), 502

    qr_code_info = primeiro_qr_code(pedido_pagbank)
    charge = (pedido_pagbank.get("charges") or [{}])[0]
    return jsonify(
        {
            "pedido_id": int(pedido["id"]),
            "payment_id": charge.get("id") or pedido_pagbank.get("id"),
            "status": charge.get("status") or "WAITING",
            "qr_code": qr_code_info.get("text"),
            "qr_code_base64": buscar_pagamento_pix_por_pedido(int(pedido["id"]))["qr_code_base64"],
            "ticket_url": extrair_link(qr_code_info, "QRCODE.PNG"),
        }
    )


@app.post("/webhook")
def webhook_pagbank():
    if not validar_assinatura_webhook():
        return {"status": "assinatura_invalida"}, 401

    corpo = request.get_json(silent=True) or {}
    order_id = str(corpo.get("id") or "").strip()
    if not order_id:
        return {"status": "ignorado"}, 200

    try:
        pedido_pagbank = consultar_pedido_pagbank(order_id)
        atualizar_status_pagamento_pagbank(pedido_pagbank)
    except RuntimeError as erro:
        return {"status": "erro_configuracao", "detalhe": str(erro)}, 500
    except Exception as erro:  # pragma: no cover - depende da API externa
        return {"status": "erro_consulta", "detalhe": str(erro)}, 500
    return {"status": "ok"}, 200


@app.get("/pedido/<int:sale_id>")
@client_required
def pedido_confirmado(sale_id):
    db = get_db()
    pedido = db.execute(
        "SELECT id, total, status, created_at, updated_at, paid_at FROM sales WHERE id = ? AND client_id = ?",
        (sale_id, session["client_id"]),
    ).fetchone()
    if not pedido:
        flash("Pedido nao encontrado.", "error")
        return redirect(url_for("loja_cliente"))
    itens = db.execute(
        """
        SELECT product_name, quantity, unit_price, total_price
        FROM sale_items
        WHERE sale_id = ?
        ORDER BY id
        """,
        (sale_id,),
    ).fetchall()
    pagamento_pix = buscar_pagamento_pix_por_pedido(sale_id)
    return render_template(
        "pedido_confirmado.html",
        pedido=pedido,
        itens=itens,
        pix_key=CONFIG["pix_key"],
        pagamento_pix=pagamento_pix,
        pagbank_configurado=pagbank_configurado(),
    )


@app.get("/pedido/<int:sale_id>/status")
@client_required
def status_pedido(sale_id):
    pedido = buscar_pedido_por_id(sale_id, client_id=session["client_id"])
    if not pedido:
        return jsonify({"erro": "Pedido nao encontrado."}), 404
    pagamento_pix = buscar_pagamento_pix_por_pedido(sale_id)
    if (
        pagamento_pix
        and pagbank_configurado()
        and str(pagamento_pix["status"] or "").upper() in {"WAITING", "AUTHORIZED"}
        and pagamento_pix["pagbank_order_id"]
    ):
        try:
            pedido_pagbank = consultar_pedido_pagbank(pagamento_pix["pagbank_order_id"])
            pedido = atualizar_status_pagamento_pagbank(pedido_pagbank) or pedido
            pagamento_pix = buscar_pagamento_pix_por_pedido(sale_id)
        except RuntimeError:
            pass
        except Exception:
            pass
    return jsonify(
        {
            "pedido_id": int(pedido["id"]),
            "status": pedido["status"],
            "total": float(pedido["total"]),
            "payment_status": pagamento_pix["status"] if pagamento_pix else None,
            "payment_id": (
                (pagamento_pix["pagbank_charge_id"] or pagamento_pix["pagbank_order_id"]) if pagamento_pix else None
            ),
            "paid_at": pedido["paid_at"],
        }
    )


@app.post("/carrinho/remover/<int:product_id>")
@client_required
def remover_carrinho(product_id):
    session["cart"] = [item for item in carrinho_atual() if item["product_id"] != product_id]
    flash("Item removido do carrinho.", "success")
    return redirect(url_for("ver_carrinho"))


@app.post("/checkout")
@client_required
def checkout():
    cart = carrinho_atual()
    if not cart:
        flash("Seu carrinho esta vazio.", "error")
        return redirect(url_for("ver_carrinho"))

    db = get_db()
    produtos = []
    total = 0.0
    for item in cart:
        produto = db.execute("SELECT * FROM products WHERE id = ?", (item["product_id"],)).fetchone()
        if not produto or produto["stock"] < item["quantity"]:
            flash(f"Estoque insuficiente para {item['name']}.", "error")
            return redirect(url_for("ver_carrinho"))
        produtos.append(produto)
        total += float(produto["price"]) * item["quantity"]

    agora = agora_texto()
    try:
        db.execute("BEGIN")
        cursor = db.execute(
            """
            INSERT INTO sales (client_id, client_name, total, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (session["client_id"], session["client_name"], total, "pendente", agora, agora),
        )
        sale_id = cursor.lastrowid
        for item, produto in zip(cart, produtos):
            novo_estoque = int(produto["stock"]) - int(item["quantity"])
            novo_status = "Disponivel" if novo_estoque > 0 else "Esgotado"
            db.execute(
                "UPDATE products SET stock = ?, status = ? WHERE id = ?",
                (novo_estoque, novo_status, produto["id"]),
            )
            db.execute(
                """
                INSERT INTO sale_items (sale_id, product_id, product_name, quantity, unit_price, total_price)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    sale_id,
                    produto["id"],
                    produto["name"],
                    item["quantity"],
                    float(produto["price"]),
                    float(produto["price"]) * item["quantity"],
                ),
            )
        db.commit()
    except sqlite3.DatabaseError:
        db.rollback()
        flash("Nao foi possivel concluir a compra.", "error")
        return redirect(url_for("ver_carrinho"))

    session["cart"] = []
    flash("Compra registrada com sucesso.", "success")
    return redirect(url_for("pedido_confirmado", sale_id=sale_id))


if __name__ == "__main__":
    host = os.getenv("FLASK_RUN_HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
    app.run(host=host, port=port, debug=debug)
